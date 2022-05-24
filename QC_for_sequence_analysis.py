from Bio import SeqIO
from Bio import AlignIO
from Bio.Align.Applications import MafftCommandline
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord
from pathlib import Path
import os
import re
import shutil
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse
import initio_utilities as iu

parser = argparse.ArgumentParser(
            prog='preparation_for_clean_up.py',
            usage='%(prog)s -d/--dir <NGS RUN location>',
            description='It takes as an argument the location of the NGS run and\
                         creates alignments of the 2-20PC sequences and the Refseq\
                         an excel file with the Quasibam file and the positions\
                         according to the fasta 20 PC and generates a report to\
                         check the frame of the sequences. It creates a directory\
                         called data_to_clean, containing the sequences to be cleaned'
                         )
parser.add_argument('--dir', '-d', required=True)
args = parser.parse_args()

# quasibams (tab and fasta)
path_to_quasibams = Path.cwd().joinpath(args.dir).joinpath('quasibams')
run_id = Path(args.dir).stem
path_to_refseq = path_to_quasibams.parents[2].joinpath(
                                            'scripts/K03455.1_HXB2.fasta')

os.chdir(path_to_quasibams)
data_to_clean = Path('data_to_clean')
tmps = Path('tmps')

if data_to_clean.exists() and tmps.exists():
#if os.path.exists(data_to_clean) and os.path.exists(tmps):
    shutil.rmtree(data_to_clean)
    shutil.rmtree(tmps)
    os.mkdir(data_to_clean)
    os.mkdir(tmps) # use just for manipulation
else:
    os.mkdir(data_to_clean)
    os.mkdir(tmps)


''' GENERATION OF AN ALIGNMENT WITH REFSEQ AND 2-20PC SEQUENCES TO USE 
    FOR THE CLEANUP'''

for tab in Path.cwd().glob('*.tabular'):
    seq_id = tab.stem
    SeqRecords = [SeqIO.read(path_to_refseq,'fasta')] 
    for f in Path(path_to_quasibams).glob(f'{seq_id}*.fas'):
        sequence = SeqIO.read(f,'fasta')
        SeqRecords.append(sequence)
    tmp_file_name = Path('tmps').joinpath(f'{seq_id}_20-2PC_HXB2.fas')
    SeqIO.write(SeqRecords, tmp_file_name, 'fasta')
    mafft_cline = MafftCommandline(input= tmp_file_name, maxiterate = 0)
    stdout, stderr = mafft_cline()
    alignment_file_name = str(tmp_file_name).replace('_20-2PC_HXB2.fas', 
                                                '_aligned_to_HXB2.fas')
    alignment_file_name = Path(alignment_file_name)

    with open(alignment_file_name, "w") as handle:
        handle.write(stdout)
   
    'create excel file with quasibam data'
    quasibam_tab = pd.read_csv(tab, sep='\t')
    #### New format QuasiBAM
    quasibam_tab = quasibam_tab.rename(columns={'Unnamed: 0' : 'Pos'})#'Sample ID'})
    ####
    fasta_positions =[]
    gap_positions = []
    #gaps = pd.DataFrame(columns=['Pos','Pos_FASTA _20%','Ref_N', 'Depth',
    #                             'A', 'C', 'G', 'T', 'Gap','Cons'])
    # New version of quasibam doesnt use column Cons
    gaps = pd.DataFrame(columns=['Pos','Pos_FASTA _20%','Ref_N', 'Depth',
                                 'A', 'C', 'G', 'T', 'Gap'])
    ins_positions = []
    insertions = pd.DataFrame(columns=['Pos','Pos_FASTA _20%','Ref_N', 
                                        'Depth', 'A', 'C', 'G', 'T', 'Ins',
                                        'I_Desc'])
    pos = 0
    for n in range(len(quasibam_tab)):
        pos += 1 
        data = quasibam_tab.iloc[n]
        if data[['I_Desc']].isnull().any() == True: 
            # d[[database]] I_Desc    NaN 
            # Name: 8916, dtype: object
            # d[database]  NaN
            fasta_positions.append(pos)
        else: 
            ins = data['I_Desc'].split(':')[0] #C:50, TA:50, << a sequence has this but 
                                                #it was N (<10 reads) Not frequent
            #percentage = data['I_Desc'].split(':')[1].replace(',','')
            percentage = data['I_Desc'].split(':')[1].split(',')[0]
            insertions = insertions.append(data[['Pos','Ref_N', 'Depth', 'A', 'C', 
                                                'G', 'T', 'Ins','I_Desc']])
            ins_positions.append(pos)
            fasta_positions.append(pos)
            if float(percentage) > 20:
                pos += len(ins)

        if float(data['Gap']) > 2:
            #gaps = gaps.append(data[['Pos','Ref_N', 'Depth', 'A', 'C', 'G', 'T', 
            #                        'Gap','Cons']])
            gaps = gaps.append(data[['Pos','Ref_N', 'Depth', 'A', 'C', 'G', 'T', 
                                    'Gap']])
            gap_positions.append(pos)

    gaps['Pos_FASTA _20%'] = gap_positions      
    insertions['Pos_FASTA _20%'] = ins_positions
    quasibam_tab['Pos_FASTA _20%'] = fasta_positions
    #quasibam_tab = quasibam_tab[['Pos','Pos_FASTA _20%','Ref_N', 'Depth', 
    #                            'A', 'C', 'G', 'T', 'Gap', 'Ins', 'I_Desc',
    #                            'Cons','qA', 'qC', 'qG', 'qT', 'Apos', 'RCod',
    #                            'RAA', 'AA Dep', 'AA','Cod']]
    quasibam_tab = quasibam_tab[['Pos','Pos_FASTA _20%','Ref_N', 'Depth', 
                                'A', 'C', 'G', 'T', 'Gap', 'Ins', 'I_Desc',
                                'Ref_AA','AA_depth', 'Cod', 'AA']]
    workbook = Workbook()
    ws = workbook.active
    ws1 = workbook.create_sheet(f'{seq_id}_FASTA_numbering', 0)
    ws2 = workbook.create_sheet(f'{seq_id}_gaps', 1)
    ws3 = workbook.create_sheet(f'{seq_id}_insertions', 2)

    for rowdf in dataframe_to_rows(quasibam_tab, index=False, header=True):
        ws1.append(rowdf)

    for rowdf in dataframe_to_rows(gaps, index=False, header=True):
        ws2.append(rowdf)

    for rowdf in dataframe_to_rows(insertions, index=False, header=True):
        ws3.append(rowdf)

    excel_file = str(tab).replace('.tabular','_Quasibam_FASTA_numbering.xlsx') 
    workbook.save(Path(excel_file))


'''GENERATION OF REPORT TO CHECK FRAMESHIFTS'''

iu.frameshift_check_sequence_locator(run_id, tmps, data_to_clean)


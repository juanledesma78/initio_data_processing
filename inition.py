__version__ = '1.0'
__date__ = '24/08/2022'
__author__ = 'Juan Ledesma'


import argparse
import pandas as pd
import numpy as np
import subprocess as sp
import os
import re
import shutil
import sys
import logging
from pathlib import Path
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord
from Bio import SeqIO
from Bio import AlignIO
from Bio.Align.Applications import MafftCommandline
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord
from Bio.Align import MultipleSeqAlignment
from Bio import pairwise2
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


program_specs = argparse.ArgumentParser(
    prog='hiv_initio_project.py',
    usage='%(prog)s --input [-pgen] [-qc] [-c] [-fg] [-j] [-pj] [-dc] [-majodc]',
    description='The script is written to perfom different functions depending\
                of the argument selected.\
                Argument --input is REQUIRED for the functions to work.\
                The other optional arguments are intended to be used one a time.'
                )

program_specs.add_argument('--input', required= True, #'-i', 
            help='Path to the run containing the data to analyse.\
                This is REQUIRED for using the tools of the script,\
                which are selected by the following arguments.\
                When using parameter -j/--jphmm this is the path\
                to the FASTA file to analyse')

program_specs.add_argument('-pgen','--post_genomancer',  #action='store_true', -jp is calling GEnomancer BAD
            nargs='?', const= 'Mm100',
            help='with input location selected (--input), decompresses the files\
                "_quasibams.zip","_post-run.zip" and "_typonomer.zip".\
                It takes the files ".tabular" stored in "quasibams", produces\
                a file "Coverage_Depth_Summary_HIV.csv" with the coverage\
                depth for all the sequences and runs the PHE QuasiBAM tool\
                to generate sequences of Majority and Minority variants\
                (frequency of 20PC and 2PC) at depth of 100 reads. It reads\
                the files ".nex" from directory "post-run" and files\
                "_read_counts.tsv" and "_genomes.fas" and creates a file\
                "Summary_Genomancer_Results_HIV.csv" with information about\
                the analysis by Genomancer. It extracts each sequence from\
                the file "_genomes.fas" into a new directory called "FASTAs".\
                [ADDITIONAL ARGUMENT] A value "majo" can be used after\
                -p/--post_genomancer to produce additional sequences at depth\
                of 30 reads for Majority variants (20PC) for the same sample.\
                <WARNING> This tool needs to be run in the HPC cluster to get\
                access to PHE QuasiBAM and to the files "_quasibams.zip"\
                ,"_post-run.zip" and "_typonomer.zip", "_genomes.fas" and\
                "_read_counts.tsv"') 
                # symbol % gives an error if included in help 

program_specs.add_argument('-qc', '--quality_control', action='store_true',
            help=' with input location selected (--input), it takes each file ".tabular"\
                from the directory "quasibams" and generates a new file\
                "_Quasibam_FASTA_numbering.xlsx", to be used for potential\
                inspection of positions in the sequences. The tool takes the\
                files ".fas" from each sample located in the same directory and\
                aligns them to the reference sequence K03455.1 HXB2, saving\
                the resulting alignmets in a subdirectory "quasibams/tmps".\
                These alignments are trimmed in sub-alignments corresponding\
                to each genomic region and evaulated for frame shifts and/or\
                stop codons. The information is recorded in the file "FRAMESHIFT\
                _initial_check_RUNID.txt" in "quasibams". In a subdirectory\
                "quasibams/data_to_clean", a file "_aligned_to_HXB2.fas" for\
                each sample is created for potential sequence editing if needed\
                , containing the query sequences aligned to all the gene segments\
                of reference.')

program_specs.add_argument('-c','--confirmation', action='store_true',
            help='with input location selected (--input), it takes the "FASTA" files\
                from the subdirectory "quasibams/data_to_clean" and repeats\
                the same evaluation for frame-shifts/stop codons, generating\
                a file "FRAMESHIFT_post_cleaning_RUNID.txt"')

program_specs.add_argument('-fg' ,'--file_generation', action='store_true',
            help='with input location selected (--input), takes the clean\
                "FASTA" files from the subdirectory "quasibams/data_to_clean" and\
                creates the files "_2-20PC_D100_seqs_for_Resistance_report.fasta"\
                and "_20PC_D100_seqs_for_WG_Subtyping.fasta" in a new directory\
                called "fastas" in the run directory, used for subtyping analysis\
                and antiviral resistance genotyping.\
                An additional file "_20PC_D30_seqs_for_All_Analyses.fasta"\
                will be created if majoritiy variants at depth of 30 reads have been\
                analysed as well.\
                New directories "hivdb.stanford.report" and "subtyping"\
                are generated to store the reports of the analyses to be done.\
                Finally a file "_2-20PC_HIV_Genome_map.fasta" is created for\
                numbering and genomic mapping purposes, containing the query\
                sequences aligned to all the gene segments of reference.')

program_specs.add_argument('-j' ,'--jphmm', action='store_true',
            help='For this tool, --input argument points the location of\
                a "FASTA" file containing the sequences to be subtyped and\
                runs jphmm locally. The fasta files should be saved in a\
                subdirectory called "fastas"\
                (i.e  /home/phe.gov.uk/user/jpHMM/fastas/INITIO/RUNID_20PC_seqs_for_WG_Subtyping.fasta )\
                The results are stored in a subdirectory "jphmm/outputs/RunID", where\
                the software has been installed')

program_specs.add_argument('-pj' ,'--post_jphmm', 
            help='This tool takes TWO arguments: the first one is the normal --input\
                that points to the RUN ID of the data analysed and the second one,\
                provided after -pj takes the path to the folders(s) containing the\
                subtyping information generated by jphmm.\
                With these two imputs, it takes the jphmm reports and creates the file\
                "Recombination_positions_NGS_Run_jpHMM.txt" and "results_NGS_Run_jpHMM.csv".\
                and saved them in "RunID/subtyping/jphmm". The ".csv" file indicates the \
                clasification of the sequences after using the tool jpHMM locally.\
                The ".txt" file shows the regions of the sequences involved in that\
                subtype clasification, very useful for predicting the breakpoints\
                in a recombination.')

program_specs.add_argument('-dc','--data_consolidation', action='store_true',
            help='with input location selected (--input), consolidates the data from\
                "RUNID_sample_list.csv", "RUNID_NGS_QC.csv", "Summary_Genomancer\
                _Results_HIV.csv", reports from subdirectory "subtyping/rega", \
                "subtyping/comet" and "subtyping/jphmm" and the stanford report\
                from "hivdb.stanford.report" generated using sequences at depth\
                of 100 reads and frequency of 20PC and 2PC (Majority and Minority\
                Variants, respectively) and creates a file "_NGS_Results_2-20PC_\
                D100.xlsx"" in the main directory of the INITIO batch.')

program_specs.add_argument('-majodc','--majority_30', action='store_true',
            help='It works similar to the argument "--data_consolidation" but it\
                is specific for those sequences generated at depth of 30 reads\
                and frequency of 20PC (Majority).')

args = program_specs.parse_args()


##############################################################################

def trimming_alignment(Bio_Aln, start_trimming_pos, end_trimming_pos):

    """It takes a Biopython alignment and returns a new alignment trimmed to
    the given starting and ending positions. Gaps (-) introduced in the refseq 
    to align it to the query sequence (insertions in query) can compromise the 
    final region to trimm. Therefore they are not considered in the numbering
    when trimming is done (if str(Bio_Aln[0].seq[pos]) != '-')
    The positions to trim can be given from a dictionary with several regions to
    trim from the same alignment.
    gene_Dictionary = {'Gag': [790, 2292], 'Pol': [2085, 5096]}
    """

    initial_pos_ungapped = 0
    final_pos_ungapped = 0
    first_pos = 0
    last_pos = 0
    for pos in range(len(Bio_Aln[0].seq)): 
        if initial_pos_ungapped <= (start_trimming_pos-1): # -1 python correction 
            if str(Bio_Aln[0].seq[pos]) != '-':
                first_pos = pos
                initial_pos_ungapped += 1
        if final_pos_ungapped <= end_trimming_pos:
            if str(Bio_Aln[0].seq[pos]) != '-':
                last_pos = pos
                final_pos_ungapped += 1
    aln = Bio_Aln[:, first_pos : last_pos]
    return aln , first_pos


##############################################################################

def frameshift_check_sequence_locator(run_ID, tmps, 
                                    report_destination, check= True ): 

    """ It reads the alignments provided in a directory "tmps", translates 
    the nt sequences at three diferent frames, aligns their amino acids to the 
    reference sequence and finds out wheter the first one (expected 
    to be in frame after trimming) has the best score. It also checks if there 
    are stop codons accross the whole sequence, except for the final stop codon.   
    It returns a dictionary with the id of the sequences, the first frame and a 
    comment if there is a frameshihf and/or stops codons, all of them used to 
    generate a report.
    Thi is an alternative to HIV SEQUENCE LOCATOR 
    https://www.hiv.lanl.gov/content/sequence/LOCATE/locate.html"""

    if check == True:
        seqloc_File_Name = f'FRAMESHIFT_initial_check_{run_ID}.txt'
        
    else:
        seqloc_File_Name = f'FRAMESHIFT_post_cleaning_{run_ID}.txt'  

    seqloc_File = open(report_destination.joinpath(seqloc_File_Name), 'w')
    info_seqcloc =''

    for aln_file in tmps.glob('*_aligned_to_HXB2.fas'):
        sample_name = str(aln_file.name).replace('_aligned_to_HXB2.fas','')
        info_seqcloc +=  f'''
------------------------------------------------------------------------------
---------------- {sample_name} ------------------------------------------------
------------------------------------------------------------------------------

'''

        # (protease: 2253-2549, RT: 2550-3869, RNase: 3870-4229, Integrase:4230-5096)
        # gp120: 6225 → 7757, gp41: 7758 → 8795

        gene_Dictionary = {'Gag': [790, 2292], 'Pol': [2085, 5096], 
                        'Vif': [5041, 5619],'Vpr': [5559, 5850], 
                        'Tat1': [5831, 6045], 'Rev1': [5970, 6045],
                        'Vpu': [6062, 6310], 'Gp160': [6225, 8795], 
                        'Tat2': [8379, 8469],'Rev2': [8379, 8653],
                        'Nef': [8797, 9388]}

        aln = AlignIO.read(aln_file, 'fasta')
        all_genes_records = [record for record in aln] #refseq and 20PC and/or 2PC

        for gene, pos in gene_Dictionary.items(): 
            # Each alignment is trimming using the coordinates of each gene  
            trimmed_aln, first_pos = trimming_alignment(aln, pos[0], pos[1]) 
                                    # 790-1, 2292)

            ''' This generates a file to map all the genes to the 2-20PC sequences.
            that can be used to check those positions involved in the shift of the 
            frame. Although the genes look aligned the file is not a proper alignment'''

            remaining_length = len(aln[0]) - (first_pos + len(trimmed_aln[0])) 
            gene_seq = ('-'*first_pos + str(trimmed_aln[0].seq) +\
                         '-'*remaining_length)
            gene_record = SeqRecord(seq= Seq(gene_seq), 
                                    id = (aln[0].id).replace('complete_genome', 
                                                            gene),
                                    description= '')
            all_genes_records.append(gene_record)
            name_file = str(aln_file).replace('tmps', 'data_to_clean')
            name_file = name_file.replace('_aligned_to_HXB2.fas', 
                                        '_all_genes_aligned.fas')
            SeqIO.write(all_genes_records, Path(name_file),'fasta')

            ''' This is the tool to check frame-shifts and stop codons'''
            # invariable, it should not have gaps, in-frame...
            AA_ref_seq = trimmed_aln[0].seq.ungap().translate() 
            NT_query_seqs = [] # stores the sequences from the same sample 
                              # trimmed for each gene
            for n in range(1, len(trimmed_aln)):
                NT_query_seqs.append(
                            SeqRecord(seq= trimmed_aln[n][:].seq.ungap(), 
                                    id= trimmed_aln[n].id, 
                                    name= '')
                                    )
            
            frame_dictionary = {} # stores the Seq ID, 1st frame and results
            for n in range(len(NT_query_seqs)):
                scores = [] # scoring for each frame
                frames = [] # to store the 3 frames and compare
                frame_analysis = [] # results
                record = NT_query_seqs[n]
                for start_pos in range(3):
                    #BiopythonWarning: Partial codon, 
                    # len(sequence) not a multiple of three.
                    codons = []
                    frame = record.seq[start_pos:]
                    for n in range(0,len(frame),3):
                        codon = str(frame[n:n+3])
                        if len(codon) == 3:
                            codons.append(codon)
                    frame = Seq(''.join(codons)).translate() # frame multiple of 3
                    frames.append(str(frame))
                    score = pairwise2.align.localxx(frame, 
                                                    AA_ref_seq, 
                                                    score_only=True)
                    scores.append(score)

                if scores[0] < scores[1] or scores[0] < scores[2]:
                    frame_analysis.append('Sequence MAY NOT BE in frame')
                
                if '*' in frames[0][:-1]:
                    frame_analysis.append('STOP CODON(s) detected along the sequence.')
               
                frame_dictionary[record.id] = [frames[0], '. '.join(frame_analysis)]
                
            #return frame_dictionary 
            for seqid, frame_info in frame_dictionary.items():
                AAseq = ''
                for n in range(len(frame_info[0])):
                    if n !=0 and n%150 == 0:
                        AAseq += '\n'
                        AAseq += frame_info[0][n]
                    else:
                        AAseq += frame_info[0][n]
                info_seqcloc += f'{seqid}\t{gene}\t{frame_info[1]}\n{AAseq}\n'
        logging.info(f'File {Path(name_file).name} generated')
    logging.info(f'Report {seqloc_File_Name} generated')
    seqloc_File.write(info_seqcloc)
    seqloc_File.close()


##############################################################################
## WORKS BUT NOT INCLUDED YET
def extracting_fastas_by_id_selection(fasta_input, ID_file):

    """It takes a given fasta file with several sequences and returns a new 
    fasta file containing the sequence IDs matching the ones given in a txt/csv
    file"""

    name_file = ID_file.stem
    with open(ID_file,'r') as seqid:
        seq_IDs = [line.rstrip() for line in seqid ]

    fasta_file = fasta_input.name
    os.chdir(Path.cwd().joinpath(fasta_input.parent))

    new_FASTA_file = [] # with duplicates
    #selection = {} # remove duplicates  
    
    for record in SeqIO.parse(fasta_file, 'fasta'): 
        contig = re.sub(r'\.(2|20)PC.*', '', record.id)
        # TRY WITH CONTIGS AND ID (NO .1)
        if contig in seq_IDs:
            sequence = SeqRecord(record.seq.ungap(), record.id, description='')
            print(f'ID:{contig} --> FASTA header:{sequence.id}')
            new_FASTA_file.append(sequence)
        
    output_file = f'{name_file}_sequences.fasta'
    SeqIO.write(new_FASTA_file, output_file,'fasta')
    print(f'''
{len(new_FASTA_file)} sequences selected and saved in "{output_file}"
''') 


##############################################################################
## TO BE TESTED
def extracting_fastas_from_file(fasta_input):

    FASTAs = Path(fasta_input.parent).joinpath(f'FASTAs')
    if FASTAs.is_dir() is False:
        os.mkdir(FASTAs)
    for record in SeqIO.parse(fasta_input, 'fasta'):
        SeqIO.write(record, FASTAs/f'{record.id}.fas', 'fasta')


##############################################################################

def parsing_paths(run_info):

    """It takes the input argument (run location) and generates some variables
    and paths needed for the other functions"""

    run_path = Path.cwd().joinpath(run_info)
    os.chdir(run_path)
    run_ID = Path.cwd().name
    batch_path = Path.cwd().parents[0]
    batch = Path.cwd().parents[0].name
    project = Path.cwd().parents[1]
    ref_seq = project.joinpath('initio_data_processing/K03455.1_HXB2.fasta') 
    quasibams = Path.cwd().joinpath('quasibams')
    data_to_clean = quasibams.joinpath('data_to_clean')
    tmps = quasibams.joinpath('tmps')

    return run_path, run_ID, batch_path, batch,\
         project, ref_seq, quasibams,\
             data_to_clean, tmps


##############################################################################

def postgenomancer(run_ID, depth, workflow='HIV'):

    """ Decompresses the files *_quasibams.zip,*_post-run.zip and *_typonomer.zip, 
    if needed, creates a dataframe with the coverage depth for all the sequences 
    using the tabular files, run QuasiBAM script to generate sequences for 
    Majority and Minority variants at depth of 100, and if needed, at depth of 30 
    reads for Majority variants. It creates an XLSX file to summarise the results
    generated by the pipeline Genomancer. I also extracts each fasta sequence from 
    the file *_genomes.fas into a new directory called FASTAs.
    ANY MODIFICATION IN GENOMANCER THAT AFECTS THE TABULAR FILES (HEADERS)
    CAN AFFECT THIS FUNCTION. 
    I.E. KeyError: "None of [Index(['I_Desc'], dtype='object')] are in the [index]"
    """

    #print(f'''
    #-- Initiating Analysis on Genomancer results -- 
    #''') 

    ''' UNZIP RESULT FILES'''

    if Path.cwd().joinpath('quasibams').exists() == False:
        quasibam = sp.run(['unzip',f'{workflow}_quasibams.zip','-d', 'quasibams'])
    if Path.cwd().joinpath('post-run').exists() == False:
        postrun = sp.run(['unzip',f'{workflow}_post-run.zip','-d', 'post-run'])
    if Path.cwd().joinpath('typonomer').exists() == False:
        typonomer = sp.run(['unzip',f'{workflow}_typonomer.zip','-d', 'typonomer'])

    
    ''' GENERATION OF DEPTH FILE USING QUASIBAM TAB FILES'''

    frequency_range = [20, 2] 
    dataframe_list =[]
    for quasibam_file in Path('quasibams').glob('*.tabular'):
        seq_id = quasibam_file.stem

        seq_id = re.sub(r'^\d+_','', seq_id) # 'H215141107.1' RS22000334.1.draft.0
        seq_id = re.sub(r'.draft.\d','', seq_id) # 'H215141107.1'
        quasi_df = pd.read_csv(quasibam_file , sep = "\t")
        quasi_df.rename(columns={'Depth': seq_id}, inplace = True)
        dataframe_list.append(quasi_df[seq_id])

        for frequency in frequency_range:
            outfile = Path('quasibams').joinpath(f'{seq_id}.{frequency}PC.fas')
            fasta_header = f'{seq_id}.{frequency}PC'
            running_quasibam_ = sp.run(["qb_post_process.pl",
                                        "-i",quasibam_file,
                                        "-o",outfile,
                                        "-s", fasta_header,
                                        "-d", "100",
                                        "-c", str(frequency),
                                        "-n", "N"])
        if depth == 'MAJO':
            outfile2 = Path('quasibams').joinpath(f'{seq_id}.20PC.D30.fas')
            fasta_header = f'{seq_id}.20PC.D30'
            running_quasibam_ = sp.run(["qb_post_process.pl",
                                    "-i",quasibam_file,
                                    "-o",outfile2,
                                    "-s", fasta_header,
                                    "-d", "30", ###
                                    "-c", "20", ###
                                    "-n", "N"])
            #    tabular_output = f'quasibams/{seq_id}.tabular'
            #    os.rename(quasibam_file, tabular_output)
            quasibam_file.rename(Path('quasibams').joinpath(f'{seq_id}.tabular'))
        else:
            pass
    Depth_dataframe = pd.concat(dataframe_list, axis = 1 )
    Depth_dataframe.index += 1
    Depth_dataframe['Position'] = Depth_dataframe.index #using the index as a column
    Depth_dataframe.insert(0, 'Position', Depth_dataframe.pop('Position'))
    Depth_dataframe.to_csv(Path.cwd().
                joinpath(f'Coverage_Depth_Summary_{workflow}.csv'), index= False)

    NGS_metrics = []
    metrics_cols = ['Sample ID', 'Length','Median Depth', 'Max', 
                    'Min',  'Depth 30', 'Depth 100', 'Contig']
    for col in Depth_dataframe.columns:
        if col != 'Position': 
            column = Depth_dataframe[col]
            coverage30reads = (column[Depth_dataframe[col]>30].
                                    count()/Depth_dataframe[col].count())*100
            coverage100reads = (column[Depth_dataframe[col]>100]
                                    .count()/Depth_dataframe[col].count())*100
            id = re.search(r'^RS\d+', col)
            id = id.group()
            print(f'Extracting data from sample {id}, contig {col}')
            NGS_metrics.append([id, Depth_dataframe[col].count(), 
                                Depth_dataframe[col].median(), 
                                Depth_dataframe[col].max(), 
                                Depth_dataframe[col].min(),
                                round(coverage30reads, 2), 
                                round(coverage100reads, 2), col
                                    ])
    new_dataframe = pd.DataFrame(NGS_metrics, columns=metrics_cols)
    
    ''' GENERATION OF SUMMARY REPORT DATAFRAME '''

    genome_dict = {
            'HCV':['Core', 'E1', 'E2', 'NS3', 'NS4b', 'NS5a', 'NS5b'],
            'HIV':['gag', 'pol', 'vif', 'vpu', 'env', 'nef'],
            'EV':['2C', '3C', '3D', 'VP1', 'VP2', 'VP3']
                    }
    #summary_dataframe = pd.read_csv(f'{workflow}_read_counts.tsv', sep = "\t")
    summary_dataframe = pd.read_csv(f'{workflow}_readcounts.tsv', sep = "\t")
    summary_dataframe = summary_dataframe.rename(columns={'Unnamed: 0' : 'FASTQ ID'})#'Sample ID'})
    summary_dataframe = summary_dataframe.rename(columns={'Raw' : 'Raw reads'})
    summary_dataframe = summary_dataframe.rename(columns={'Trimmed' : 'Trimmed reads'})
    summary_dataframe = summary_dataframe.rename(columns={'Post-prinseq' : 'Post-prinseq reads'})
    summary_dataframe = summary_dataframe.rename(columns={'Dehumanised' : 'Dehumanised reads'})
    #summary_dataframe = summary_dataframe.rename(columns={'Mapped' : 'Mapped reads'})
    summary_dataframe = summary_dataframe.rename(columns={'viromancer' : 'Mapped reads'})
    summary_dataframe['% Mapped reads'] = round(
                        (summary_dataframe['Mapped reads']*100)/summary_dataframe['Raw reads'], 
                        2)
    summary_dataframe['Fasta sequences by the pipeline'] = ''

    run_sample_list = []
    for fastq in summary_dataframe['FASTQ ID']:
        seq_id = re.sub(r'^\d+_','', fastq)
        run_sample_list.append(seq_id)
    summary_dataframe['Sample ID'] = run_sample_list #'H220141691'
    summary_dataframe.insert(0, 'Sample ID', summary_dataframe.pop('Sample ID'))
    summary_dataframe['MiSeq run'] = run_ID
    summary_dataframe.insert(1, 'MiSeq run', summary_dataframe.pop('MiSeq run'))
    fastq_list = list(summary_dataframe['FASTQ ID']) # '1574517_H220141691'

    if workflow in genome_dict.keys():
        for gene in genome_dict[workflow]:
            summary_dataframe[gene] = ""

    ''' IDENTIFYING THE FASTA SEQUENCES GENERATED BY GENOMANCER '''

    fasta_sequence_results = [] 
    # RS22000298.1 ; RS22000298.1 ; RS22000298.1 ; RS22000298.1 ; RS22000298.1 ; RS22000298.1
    # set() to remove duplicates???

    with open(f'{workflow}_genomes.fas','r') as genfas:
        fasta_file = genfas.read()
        for fastq in fastq_list:
            #sample = re.sub(r'^\d+_','', sample)
            pattern = f'{fastq}\.\d+'
            sequence_ids = re.findall(pattern, fasta_file)
            info =""
            for genome in sequence_ids:
                    if len(info)>0:
                        info += " ; " + genome
                    else:
                        info += genome
            fasta_sequence_results.append(info)

    summary_dataframe['Fasta sequences by the pipeline'] = fasta_sequence_results

    FASTAs = Path.cwd().joinpath(f'FASTAs')
    if FASTAs.is_dir() is False:
        os.mkdir(FASTAs)
    for record in SeqIO.parse(Path.cwd().
                            joinpath(f'{workflow}_genomes.fas'), 'fasta'):
        SeqIO.write(record, FASTAs/f'{record.id}.fas', 'fasta')

    
    ''' FINDING OUT WHETHER THE SEQUENCES ARE INCLUDED IN THE TREES '''
    
    for phylo_file in Path('post-run').glob('*.nex'):
        with open(phylo_file, 'r') as newick:
            gene = (newick.name).replace('post-run/','').replace('.nex','')
            tree = newick.read()
            for n in range(len(summary_dataframe)):
                row_info = ''
                genomes = summary_dataframe.loc[n]['Fasta sequences by the pipeline'].split(' ; ')
                for i in range(len(genomes)):
                    if genomes[i] in tree:
                        if len(row_info)>0:
                            row_info += " ; " + genomes[i]
                        else:
                            row_info += genomes[i]

                summary_dataframe.loc[n, [gene]] = row_info
        #phylo_file.rename(Path('post-run').joinpath(f'{runid}_{phylo_file.name}'))

    summary_dataframe = pd.merge(summary_dataframe, new_dataframe, how= 'left', on='Sample ID')
    summary_dataframe.insert(9, 'Depth 100', summary_dataframe.pop('Depth 100'))
    summary_dataframe.insert(9, 'Depth 30', summary_dataframe.pop('Depth 30'))
    summary_dataframe.insert(9, 'Min', summary_dataframe.pop('Min'))
    summary_dataframe.insert(9, 'Max', summary_dataframe.pop('Max'))
    summary_dataframe.insert(9, 'Median Depth', summary_dataframe.pop('Median Depth'))
    summary_dataframe.insert(9, 'Length', summary_dataframe.pop('Length'))
    summary_dataframe.insert(9, 'Contig', summary_dataframe.pop('Contig'))
    # summary_dataframe.replace('', 'N/A', inplace = True)
    # summary_dataframe.replace('NaN', 'N/A', inplace = True)
    summary_dataframe.to_csv(Path.cwd().
                joinpath(f'Summary_Genomancer_Results_{workflow}.csv'), index= False)

    #print(f'''
    #-- Process successfully COMPLETE -- 
    #''')  

##############################################################################

def qc_for_sequence_analysis(quasibams, ref_seq, data_to_clean, tmps):

    """" Aligns a reference sequence to the nucloetide sequences generated from
    the sample and save the alignmet in "data_to_clean". The alignment can be 
    used to clean the sequences if needed.
    TXT report with amino acids sequences, potential frame shifts or stop codons
    present in the sequences is created and saved in "quasibams".
    An XLSX file containing the Quasibam file with an extra column with the 
    approximate positions according to the fasta 20 PC is created."""
    
    #print(f'''
    #-- Initiating Quality Control (QC) for sequences -- 
    #''') 
    
    if data_to_clean.exists() and tmps.exists():
        shutil.rmtree(data_to_clean)
        shutil.rmtree(tmps)
        os.mkdir(data_to_clean)
        os.mkdir(tmps) # use just for manipulation
    else:
        os.mkdir(data_to_clean)
        os.mkdir(tmps)

    ''' GENERATION OF AN ALIGNMENT WITH REFSEQ AND 2-20PC SEQUENCES TO USE 
    FOR THE CLEANUP'''
    os.chdir(quasibams)
    for tab in Path.cwd().glob('*.tabular'):
        logging.info(f'Processing quasibam file {tab.name}')
        seq_ID = tab.stem
        seq_records = [SeqIO.read(ref_seq,'fasta')] 
        for f in Path(quasibams).glob(f'{seq_ID}*.fas'):
            sequence = SeqIO.read(f,'fasta')
            seq_records.append(sequence)
            
        tmp_file_name = Path('tmps').joinpath(f'{seq_ID}_20-2PC_HXB2.fas')
        SeqIO.write(seq_records, tmp_file_name, 'fasta')
        mafft_cline = MafftCommandline(input= tmp_file_name, maxiterate = 0)
        stdout, stderr = mafft_cline()
        alignment_file_name = str(tmp_file_name).replace('_20-2PC_HXB2.fas', 
                                                    '_aligned_to_HXB2.fas')
        alignment_file_name = Path(alignment_file_name)
        
        with open(alignment_file_name, "w") as handle:
            handle.write(stdout)
            logging.info(f'Generating temporary alignment {alignment_file_name.name}')

        'create excel file with quasibam data'
        quasibam_df = pd.read_csv(tab, sep='\t')
        #### New format QuasiBAM
        quasibam_df = quasibam_df.rename(columns={'Unnamed: 0' : 'Pos'})#'Sample ID'})
        
        fasta_positions =[]
        gap_positions = []
        #gap_df = pd.DataFrame(columns=['Pos','Pos_FASTA _20%','Ref_N', 'Depth',
        #                             'A', 'C', 'G', 'T', 'Gap','Cons'])
        # New version of quasibam doesnt use column Cons
        gap_df = pd.DataFrame(columns=['Pos','Pos_FASTA _20%','Ref_N', 'Depth',
                                     'A', 'C', 'G', 'T', 'Gap'])
        insertion_positions = []
        insertion_df = pd.DataFrame(columns=['Pos','Pos_FASTA _20%','Ref_N', 
                                            'Depth', 'A', 'C', 'G', 'T', 'Ins',
                                            'I_Desc'])
        pos = 0
        for n in range(len(quasibam_df)):
            pos += 1 
            data = quasibam_df.iloc[n]
            if data[['I_Desc']].isnull().any() == True: 
                # d[[database]] I_Desc    NaN 
                # Name: 8916, dtype: object
                # d[database]  NaN
                fasta_positions.append(pos)
            else: 
                ins = data['I_Desc'].split(':')[0] 
                #C:50, TA:50, << a sequence has this but 
                #it was N (<10 reads) Not frequent
                #percentage = data['I_Desc'].split(':')[1].replace(',','')
                percentage = data['I_Desc'].split(':')[1].split(',')[0]
                insertion_df = insertion_df.append(data[['Pos','Ref_N', 
                                                        'Depth', 'A', 'C', 'G',
                                                        'T', 'Ins','I_Desc']])
                insertion_positions.append(pos)
                fasta_positions.append(pos)
                if float(percentage) > 20:
                    pos += len(ins)

            if float(data['Gap']) > 2:
                #gap_df = gap_df.append(data[['Pos','Ref_N', 'Depth', 
                #                           'A', 'C', 'G', 'T', 
                #                           'Gap','Cons']])
                gap_df = gap_df.append(data[['Pos','Ref_N', 'Depth', 
                                            'A', 'C', 'G', 'T', 
                                            'Gap']])
                gap_positions.append(pos)

        gap_df['Pos_FASTA _20%'] = gap_positions      
        insertion_df['Pos_FASTA _20%'] = insertion_positions
        quasibam_df['Pos_FASTA _20%'] = fasta_positions
        #quasibam_df = quasibam_df[['Pos','Pos_FASTA _20%','Ref_N', 'Depth', 
        #                            'A', 'C', 'G', 'T', 'Gap', 'Ins', 'I_Desc',
        #                            'Cons','qA', 'qC', 'qG', 'qT', 'Apos', 'RCod',
        #                            'RAA', 'AA Dep', 'AA','Cod']]
        quasibam_df = quasibam_df[['Pos','Pos_FASTA _20%','Ref_N', 'Depth', 
                                    'A', 'C', 'G', 'T', 'Gap', 'Ins', 'I_Desc',
                                    'Ref_AA','AA_depth', 'Cod', 'AA']]
        workbook = Workbook()
        ws = workbook.active
        ws1 = workbook.create_sheet(f'{seq_ID}_FASTA_numbering', 0)
        ws2 = workbook.create_sheet(f'{seq_ID}_gaps', 1)
        ws3 = workbook.create_sheet(f'{seq_ID}_insertions', 2)

        for rowdf in dataframe_to_rows(quasibam_df, index=False, header=True):
            ws1.append(rowdf)

        for rowdf in dataframe_to_rows(gap_df, index=False, header=True):
            ws2.append(rowdf)

        for rowdf in dataframe_to_rows(insertion_df, index=False, header=True):
            ws3.append(rowdf)

        excel_file = str(tab).replace('.tabular','_Quasibam_FASTA_numbering.xlsx') 
        workbook.save(Path(excel_file))

    '''GENERATION OF REPORT TO CHECK FRAMESHIFTS'''
    #iu.frameshift_check_sequence_locator(run_ID, tmps, quasibams)
    frameshift_check_sequence_locator(run_ID, tmps, quasibams)
    #print(f'''
    #-- QC successfully COMPLETE -- 
    #''')    


##############################################################################

def confirmation_of_sequence_cleaning(run_ID, ref_seq, quasibams, 
                                        data_to_clean, tmps):

    """Creates a second TXT report in "quasibams" to confirm the frames and the
    presence of stop codons in the sequences after modification."""
    
    os.chdir(quasibams)
    for fasta in data_to_clean.glob('*.fas'):
        frequency_check = [SeqIO.read(ref_seq,'fasta')] 
        seq_records = SeqIO.parse(fasta, 'fasta')
        for record in seq_records:
            if '.20PC' in record.id or '.2PC' in record.id :
                record.name = ''
                record.description = ''
                record.seq = record.seq.ungap('-').upper()
                frequency_check.append(record)
        tmp_file_name = str(fasta).replace('all_genes_aligned.fas', 
                                            'aligned_to_HXB2.fas')
        tmp_file_name = tmp_file_name.replace('data_to_clean','tmps')
        SeqIO.write(frequency_check, tmp_file_name , 'fasta')
        mafft_cline = MafftCommandline(input= tmp_file_name, maxiterate = 0)
        stdout, stderr = mafft_cline()
        with open(tmp_file_name, "w") as handle:
            handle.write(stdout)
            logging.info(f'Generating temporary alignment {Path(tmp_file_name).name}')

    frameshift_check_sequence_locator(run_ID, tmps, quasibams, check=False)
    

##############################################################################

def file_generation_for_data_analysis(run_ID, ref_seq, quasibams, 
                                        data_to_clean, tmps):

    """" It takes the clean fasta files (2PC, 20PC) and creates the files needed 
    to perform subtyping and antiviral resistance genotyping in a directory called
     "fastas" in the run directory.
    It also craetes the folders hivdb.stanford.report and subtyping where to save 
    the results after doing the data analysis"""

    # current working directory is run path
    antiviral_all_freqs_depth100 = []
    subtyping_majority_depth100 = []
    majority_depth30_data = []

    for fasta in data_to_clean.glob('*_all_genes_aligned.fas'):
        temp_genome_map_file = [SeqIO.read(ref_seq,'fasta')] 
        seq_records = SeqIO.parse(fasta, 'fasta')
        for record in seq_records:
            if '.20PC' in record.id:
                record.name = ''
                record.description = ''
                record.seq = record.seq.ungap('-').upper()
                if '.D30' in record.id:
                    majority_depth30_data.append(record)
                    temp_genome_map_file.append(record)
                else:
                    subtyping_majority_depth100.append(record)
                    antiviral_all_freqs_depth100.append(record)
                    temp_genome_map_file.append(record)
            if '.2PC' in record.id:
                record.name = ''
                record.description = ''
                record.seq = record.seq.ungap('-').upper()
                antiviral_all_freqs_depth100.append(record)
                temp_genome_map_file.append(record)

        tmp_file_name = str(fasta).replace('_all_genes_aligned.fas', 
                                      '_2-20PC_HIV_Genome_map.fasta')
        tmp_file_name = tmp_file_name.replace('data_to_clean','tmps')
        SeqIO.write(temp_genome_map_file, tmp_file_name , 'fasta')
        mafft_cline = MafftCommandline(input= tmp_file_name, maxiterate = 0)
        stdout, stderr = mafft_cline()
        with open(tmp_file_name, "w") as handle:
            handle.write(stdout)

        """create the template to map the HIV genome, useful for mutations
         in Polymerase"""

        genome_coordinates = {
            'Gag': [790, 2292], 'Polymerase': [2085, 5096], 
            'Protease': [2253, 2549],'RT-RNase': [2550, 4229], 
            'RT': [2550, 3869], 'RNase': [3870, 4229], 
            'Integrase': [4230, 5096], 'Vif': [5041, 5619], 
            'Vpr': [5559, 5850], 'Tat1': [5831, 6045], 'Rev1': [5970, 6045], 
            'Vpu': [6062, 6310], 'Gp160': [6225, 8795], 'Tat2': [8379, 8469], 
            'Rev2': [8379, 8653],'Nef': [8797, 9388]
            }
        aln = AlignIO.read(tmp_file_name, 'fasta')
        all_genes_records = [record for record in aln]
        for gene, pos in genome_coordinates.items():   
            trimmed_aln, first_pos = trimming_alignment(aln, 
                                                        pos[0], 
                                                        pos[1]) 
            remaining_length = len(aln[0]) -\
                             (first_pos + len(trimmed_aln[0])) 
            gene_seq = ('-'*first_pos + str(trimmed_aln[0].seq) +\
                         '-'*remaining_length)
            gene_record = SeqRecord(
                            seq= Seq(gene_seq), 
                            id = (aln[0].id).replace('complete_genome', gene),
                            description= ''
                            )
            all_genes_records.append(gene_record)
            name_file = tmp_file_name.replace('tmps', '') 
            SeqIO.write(all_genes_records, Path(name_file),'fasta')

    
    if (Path('fastas').exists() and 
        Path('hivdb.stanford.report').exists() and 
        Path('subtyping').exists()):
        pass
    else:
        os.mkdir('fastas')
        os.mkdir('hivdb.stanford.report')
        os.makedirs('subtyping/comet')
        os.makedirs('subtyping/jphmm')
        os.makedirs('subtyping/rega')
        os.makedirs('subtyping/rip')
        logging.info('Creation of subdirectories fastas, hivdb.stanford.report and subtyping ')

    resistance_file_name = Path.cwd().joinpath(
                f'fastas/{run_ID}_2-20PC_D100_seqs_for_Resistance_report.fasta'
                )
    subtyping_file_name = Path.cwd().joinpath(  
                f'fastas/{run_ID}_20PC_D100_seqs_for_WG_Subtyping.fasta'
                ) 
    SeqIO.write(antiviral_all_freqs_depth100, resistance_file_name,'fasta')
    SeqIO.write(subtyping_majority_depth100, subtyping_file_name , 'fasta')
    logging.info(f'File {resistance_file_name.name} generated')
    logging.info(f'File {subtyping_file_name.name} generated')

    if len(majority_depth30_data) > 0:
        majority_depth30_File = Path.cwd().joinpath(  
                    f'fastas/{run_ID}_20PC_D30_seqs_for_All_Analyses.fasta') 
        SeqIO.write(majority_depth30_data, majority_depth30_File , 'fasta')
        logging.info(f'File {majority_depth30_File.name} generated')

    logging.info(f'Files are available in {run_ID}/fastas')
    '''ORIGINAL 2 AND 20 PC FASTA NOT NEEDED ANYMORE'''
    #temporary_files = str(data_to_clean).replace('data_to_clean','tmps')
    shutil.rmtree(Path(tmps))
    shutil.rmtree(Path(data_to_clean))
    for frequency_fasta in quasibams.glob('*PC*.fas'):
        os.remove(frequency_fasta) # 2-20PC
    

##############################################################################
def running_local_jphmm(run_info):
    

    fasta_file_path = Path(run_info)
    run_id = fasta_file_path.stem
    outputs = Path(str(fasta_file_path.parent).\
                        replace('fastas', 'outputs')).joinpath(run_id)

    run_parts = str(run_info).split('/fastas/')
    local_jpHMM = Path(run_parts[0])
    inputs = Path('fastas').joinpath(run_parts[1])

    os.chdir(local_jpHMM)

    for fasta_seq in SeqIO.parse(inputs, 'fasta'): 
        logging.info(f'Reading sequence {fasta_seq.id}')

    try:

        os.makedirs(outputs) # add exception if they exist
        run_jphmm = sp.run(['src/jpHMM','-s', f'{inputs}', '-o', f'{outputs}', 
            '-v', 'HIV', 
            '-P', 'priors' ,
            '-I' ,'input', 
            '-Q', 'blat'], #text= True, capture_output=True)
            )#stdout=sp.PIPE, stderr=sp.PIPE )
        ##############################################
        ##   It needs to show the ID of the sequences analysed
        #  and potential Process killed
        #  for the user to re-run the remaining sequences 
        ###############################################
        #print(run_jphmm.stdout)
        #print(run_jphmm.stderr)
        ###############################################
        #if run_jphmm.returncode == 0:
        #    print(f'Process successfully performed')
        #else:
        #    print("Something went wrong")



        ##############################################

    except FileExistsError: 
        logging.warning(f'''
    < DIRECTORY ERROR>
    The directory "{outputs}" already exists.
    The run may have been analysed before. 
    ''')    
        sys.exit(1)


##############################################################################
def process_jphmm_results(jphmm_report, run_ID):

    
    #jphmm_report = args.postjphmm
    jphmm_destination = Path.cwd().joinpath('subtyping/jphmm')

    jphmm_results_path = Path.cwd().joinpath(jphmm_report)
    jphmm_run_ID = ''.join(re.findall(r'Run\d*_NGS\d*', jphmm_results_path.stem))
    logging.info(f'Analysing jpHMM results for {jphmm_run_ID}')
    
    if jphmm_run_ID == run_ID:
    
        os.chdir(jphmm_results_path.parent)
        subtype_data = []
        position_data = []
        for folder in Path.cwd().glob(f'{run_ID}*'):
            logging.info(f'Data from directory {folder.name}')
            subtyping_summary = pd.DataFrame(columns=['Sample ID','Sequence ID', 
                                                    'Other', 'Subtype'])
            with open(f'{folder}/recombination_without_positions.txt','r') as report:

                lines = report.readlines()
                for line in lines:
                    if '>' in line:
                        info = line.rstrip().replace('>','').split('\t')
                        depth =''
                        if 'D30' in info[0]:
                            depth = '_D30'
                        sample_ID = re.sub(r'_[0-9](_|.)(20|2)PC(.D30)*' ,'' ,info[0]) ### identify the D30
                        fasta_header =info[0].replace('_','.')
                        info.insert(0, sample_ID)
                        logging.info(f'SAMPLE id: {sample_ID}, FASTA id: {fasta_header}')
                        info.pop(2) # returns the subtype
                        subtype_data.append(info)

            with open(f'{folder}/recombination.txt','r') as report:
                lines = report.readlines()
                for line in lines:
                    if '#' not in line:
                        if '>' in line:
                            line = line.replace('_', '.')
                            line = re.sub(r'\(bw=\de-\d+\)', '', line)    
                        position_data.append(line)


        subtyping_summary = pd.DataFrame(subtype_data, columns=['Id','Sequence', 'JPHMM'])
        subtyping_summary['Sequence'] = subtyping_summary['Sequence'].str.replace('_','.') 
        '''destination'''
        os.chdir(jphmm_destination)

        subtyping_summary.to_csv(f'results_{run_ID}{depth}_jpHMM.csv', index= False)
        with open(f'Recombination_positions_{run_ID}{depth}_jpHMM.txt', 'w') as recombinants:
            recombinants.write(''.join(position_data))

    
    else:
        logging.warning(f'''
    {run_ID} to be analysed does not match the jpHMM report id {jphmm_run_ID}
        ''')



##############################################################################
def sample_info(run_ID):

    """ It produces a dataframe sample_list by combining two files: 
    RUNXXX_sample_list.csv and RUNXXX__NGS_QC.csv.
    RUNXXX__NGS_QC.csv is created from the worksheet tab "Assign libraries to 
    pools" of "VW1922.11_Calculator_For_Pooling_Libraries_*.xlsm", which has 
    had different formats affecting the column headers. If the worksheet is 
    changed again and affects the headers this should be considered to avoid 
    errors in the csv when this function is executed."""

    try: 
        sample_list = pd.read_csv(Path.cwd().\
                                joinpath(f'{run_ID}_sample_list.csv'))
        sample_list['NGS Run'] = run_ID
        
        #QC_df = pd.read_csv(Path.cwd().joinpath(f'{run_ID}_NGS_QC.csv'))
        QC_df = pd.read_csv(Path.cwd().joinpath(f'{run_ID}_NGS_QC.csv'),\
            encoding= 'unicode_escape') # in case of error include)
        
        cols = ['Molis No.', 'RT-qPCR Ct in the extract', 
                'qPCR Ct in the library', 'Glomax', 'Library added']
        
        selected_cols = []
        for dc in cols:
            for qc_cols in QC_df.columns:
                if dc in qc_cols:
                    selected_cols.append(qc_cols)
    
        QC_df = QC_df[selected_cols]
        #QC_df.rename(columns={QC_df.columns[0]:'Sample ID'}, inplace=True)
        QC_df.rename(columns={QC_df.columns[1]:'RT-qPCR Ct'}, inplace=True)
        QC_df.rename(columns={QC_df.columns[2]:'Lib qPCR Ct'}, inplace=True)
        QC_df.rename(columns={QC_df.columns[3]:'Total DNA (ng/µl)'}, inplace=True)
        QC_df.rename(columns={QC_df.columns[4]:'Pool Number'}, inplace=True)
        
        sample_list = sample_list.merge(QC_df, how='left', left_on='RS ID',
                                                         right_on='Molis No.')
        sample_list['NGS processing output'] = ''
        sample_list = sample_list[['RS ID', 'NGS Run', 'Period','Original VL',
                    'RT-qPCR Ct','Lib qPCR Ct', 'Total DNA (ng/µl)',
                    'Pool Number','NGS processing output']]
        logging.info('Sample data added') 
        return sample_list
    
    except FileNotFoundError: 
        logging.warning(f'''
    < SAMPLE INFO FILE ERROR>
    Please Confirm that the files 
    "{run_ID}_sample_list.csv" and "{run_ID}_NGS_QC.csv" 
    exist in the folder "{run_ID}".
    ''')    
        sys.exit(1)
    

##############################################################################
def sequencing_metrics(sample_list):

    """ It produces a dataframe using the Sequencing metrics obtained after 
    executing the function postgenomancer on the data generated by the pipeline
    Genomancer. The data matching the Sample ID from the sample_list is
    added to the initial dataframe. 
    """

    try:
        NGSmetrics = pd.read_csv('Summary_Genomancer_Results_HIV.csv')
        NGSmetrics.rename(columns={'Fasta sequences by the pipeline': 
                                    'Fasta generated'}, inplace= True)
        NGSmetrics['Fasta generated']= NGSmetrics['Fasta generated'].\
                                    str.replace(r'\d*_', '' , regex= True)
        NGSmetrics = NGSmetrics[['Sample ID', 'MiSeq run', 'Raw reads', 
                                'Mapped reads','% Mapped reads','Depth 30', 
                                'Depth 100','Median Depth', 'Max', 'Min',
                                'Length', 'Fasta generated', 'Contig'
                                ]]
        '''Grouping Depths according to >90, 70-90, 50-70, <50 '''
        def categorise(row, n):    
            if row[f'Depth {n}'] > 95:
                return '>95%'
            elif row[f'Depth {n}'] > 85 and row[f'Depth {n}'] <= 95:
                return '85-95%'
            elif row[f'Depth {n}'] > 75 and row[f'Depth {n}'] <= 85:
                return '75-85%'
            elif row[f'Depth {n}'] > 50 and row[f'Depth {n}'] <= 75:
                return '50-75%'
            elif row[f'Depth {n}'] < 50:
                return '<50%'
            return np.nan
        
        depth = [30, 100]
        for d in depth:
            NGSmetrics[f'Coverage {d}'] = NGSmetrics.\
                                apply(lambda row: categorise(row, d), axis=1)
        
        NGSmetrics.insert(7, 'Coverage 100',NGSmetrics.pop('Coverage 100'))
        NGSmetrics.insert(7, 'Coverage 30',NGSmetrics.pop('Coverage 30'))
        
        sample_list = sample_list.merge(NGSmetrics, how='left', 
                                        left_on='RS ID',
                                        right_on='Sample ID')
        sample_list.drop('Sample ID', axis=1, inplace=True)
        logging.info('Sequencing metrics information added')
        return sample_list
        
    
    except FileNotFoundError: 
        logging.warning(f'''
    < GENOMANCER SUMMARY FILE ERROR >
    The file "Summary_Genomancer_Results_HIV.csv" does NOT exist
    in the folder "{run_ID}". 
    ''') 
        sys.exit(1)    


##############################################################################

def subtyping_data(run_ID, sample_list, Depth=''):

    ''' Filters the information of the subtyping reports creating a new dataframe
    and adds the new information matching the Contig Id to the sample_list dataframe. 
    '''

    try:
        rega_results = Path.cwd().joinpath('subtyping/rega')
        rega_df = pd.read_csv(rega_results.joinpath(f'results_{run_ID}{Depth}.csv'))
        rega_df['Contig ID'] = rega_df['name'].\
                            str.replace(r'\.(2|20)PC.*', '' , regex= True)
        rega_df['REGA'] = rega_df['assignment'].\
                            str.replace('HIV-1 Subtype ','').\
                            str.replace('HIV-1 ','').\
                            str.replace('Recombinant of ','').\
                            str.replace('CRF ','')
    
        rega_df = rega_df[['Contig ID', 'REGA']]
        sample_list = sample_list.merge(rega_df, how='left', 
                                        left_on='Contig',
                                        right_on='Contig ID')
        sample_list.drop('Contig ID', axis=1, inplace=True) 
        # name is sequence header
        logging.info('REGA subtyping data added')
    
    except FileNotFoundError: 
        logging.warning(f'''
    < REGA SUMMARY FILE ERROR>
    There is no file called "results_{run_ID}{Depth}.csv" in "{run_ID}/subtyping/rega".
    Confirm the file has been labelled according to the convention 
    "results_runXX_NGSXXX.csv".
    ''')
        sys.exit(1)    
    
    ''' COMET subtyping'''
    #pandas.errors.ParserError: Error tokenizing data. C error: 
    # Expected 1 fields in line 4, saw 2 for some sequences the subtyping 
    # is weird a affect the csv file and the tabs
    try:
        comet_results = Path.cwd().joinpath('subtyping/comet')
        #any manual modification in the COMET file, saved as csv file will affect this line
        comet_df = pd.read_csv(comet_results.joinpath(f'results_{run_ID}{Depth}.csv'), 
                                sep='\t')
        comet_df['Contig ID'] = comet_df['name'].\
                                str.replace(r'\.(2|20)PC.*', '' , regex= True)
        comet_df.rename(columns={'subtype':'COMET'}, inplace = True)
        comet_df = comet_df[['Contig ID','COMET']]
        sample_list = sample_list.merge(comet_df, how='left', 
                                        left_on='Contig',
                                        right_on='Contig ID')
        sample_list.drop('Contig ID', axis=1, inplace=True)
        logging.info('COMET subtyping data added') 
        
    
    except FileNotFoundError: 
        logging.warning(f'''
    < COMET SUMMARY FILE ERROR >
    There is no file called "results_{run_ID}{Depth}.csv" in "{run_ID}/subtyping/comet".
    Confirm the file has been labelled according to the convention 
    "results_runXX_NGSXXX.csv".
    ''')
        sys.exit(1)
    
    ''' JPHMM subtyping'''
    try:
        jphmm_results = Path.cwd().joinpath('subtyping/jphmm')
        jphmm_df = pd.read_csv(jphmm_results.\
                                joinpath(f'results_{run_ID}{Depth}_jpHMM.csv'))
        jphmm_df['Contig ID'] = jphmm_df['Sequence'].\
                                str.replace(r'\.(2|20)PC.*', '' , regex= True)
        jphmm_df = jphmm_df[['Contig ID', 'JPHMM']]
        sample_list = sample_list.merge(jphmm_df, how='left',
                                        left_on='Contig',
                                        right_on='Contig ID')
        sample_list.drop('Contig ID', axis=1, inplace=True)
        logging.info('JPHMM subtyping data added') 
        
    
    except FileNotFoundError: 
        logging.warning(f'''
    < JPHMM SUMMARY FILE ERROR>
    There is no file called "results_{run_ID}{Depth}_jpHMM.csv" in "{run_ID}/subtyping/jphmm".
    Confirm the file has been labelled according to the convention 
    "results_RunXX_NGSXXX{Depth}_jpHMM.csv".
    ''')
        sys.exit(1)
    
    sample_list['SCUEAL'] = '' # just in case we want to add it later on for WG
    return sample_list

################################################################

def fasta_batches(run_ID, sample_list, depth= 100):
    
    if depth == 100:
        report_file = '2-20PC_D100_seqs_for_Resistance_report.fasta'
        batch_file = 'sequences_2-20_PC_Depth_100.fasta'

    if depth == 30:
        report_file = '20PC_D30_seqs_for_All_Analyses.fasta'
        batch_file = 'sequences_Majority_Variants_Depth_30.fasta'

    # create fasta batches
    batches = sample_list.drop_duplicates(subset='Period')
    batches = batches['Period'].tolist()

    if Path.cwd().joinpath('FASTA_sequences').exists():
        pass
    else:
        logging.info('Creating folder FASTA_sequences')
        os.mkdir(Path.cwd().joinpath('FASTA_sequences'))
     
    for period in batches:
        new_sequences =[]        
        period_contigs = sample_list[sample_list['Period']== period]
        contigs = period_contigs['Contig'].dropna().tolist()

        for seq in SeqIO.parse(Path.cwd().joinpath(
            f'{run_ID}/fastas/{run_ID}_{report_file}'), 
            'fasta'):
            for contig in contigs:           
                if contig in seq.id:
                    #print(contig, f'Period_{period}', seq.id)
                    logging.info(f'Period: {period}, Contig: {contig} , Sequence: {seq.id}')
                    new_sequences.append(seq)
        
        period_fasta_file = Path.cwd().joinpath(
            f'FASTA_sequences/INITIO_{period}_{batch_file}')
        if period_fasta_file.exists():
            period_file = SeqIO.parse(Path.cwd().joinpath(period_fasta_file), 'fasta')
            records = [seq for seq in period_file]
            for new in new_sequences:
                records.append(new)
            logging.info(f'Updating file {period_fasta_file.name} with {len(new_sequences)} new records.')
            SeqIO.write(records, period_fasta_file, 'fasta')
        else:       
            logging.info(f'Updating file {period_fasta_file.name} with {len(new_sequences)} new records.')
            SeqIO.write(new_sequences, period_fasta_file, 'fasta')
            


##############################################################################
 
def data_consolidator(run_ID, batch, batch_path):

    """ Consolidates all the sample information about the subtyping and resistance
    genotyping using sequences at depth of 100 reads and frequency of 20PC and 2PC 
    (Majority and Minority Variants, respectively) in an XLSX file in the main 
    directory of the INITIO batch."""

    sample_list = sample_info(run_ID)
    sample_list = sequencing_metrics(sample_list)
    sample_list = subtyping_data(run_ID, sample_list)

    
    ''' Antiviral resistance genotyping, from July there are new changes in the
    reports, so it is need to accomodate these to take the info from the old and 
    new features'''
    
    stanford_results = Path.cwd().joinpath('hivdb.stanford.report')
    os.chdir(stanford_results)

    if stanford_results.joinpath(f'{run_ID}.zip').is_file(): 

        if Path.cwd().joinpath(run_ID).exists() == False: ##############: #

            sp.run(['unzip','-j', f'{run_ID}.zip', '-d', f'{run_ID}'])#, text=True)

            try:

                SequenceSummary = pd.read_csv(
                                Path(f'{run_ID}').joinpath('SequenceSummary.tsv'), 
                                sep = '\t')
                ResistanceSummary = pd.read_csv(
                                Path(f'{run_ID}').joinpath('ResistanceSummary.tsv')
                                , sep = '\t')
                csv = { 'SequenceSummary': SequenceSummary, 
                        'ResistanceSummary': ResistanceSummary}
                Pcnt_Mix = 'Pcnt Mix'
                StopCodons = 'StopCodons'
                Num_BDHVN = 'Num BDHVN'
                BDHVN = 'BDHVN' 
                UnusualMutations = 'UnusualMutations'
                
            except:

                SequenceSummary = pd.read_csv(
                                Path(f'{run_ID}').joinpath('sequenceSummaries.csv'))
                ResistanceSummary = pd.read_csv(
                                Path(f'{run_ID}').joinpath('resistanceSummaries.csv'))
                csv = { 'sequenceSummaries': SequenceSummary, 
                        'resistanceSummaries': ResistanceSummary}
                Pcnt_Mix = 'NA Mixture Rate (%)'
                StopCodons = 'Stop Codons'
                Num_BDHVN = 'Num Ambiguous'
                BDHVN = 'Ambiguous'
                UnusualMutations = 'Unusual Mutations'
                
            MutSummary = pd.DataFrame()
            DrugScoreSummary = pd.DataFrame()
            for title, report in csv.items():
                report['Contig ID'] = report['Sequence Name'].\
                                        str.replace(r'\.(2|20)PC.*', '' , 
                                        regex= True)
                report['Sample ID'] = report['Sequence Name'].\
                                        str.split('.').\
                                        str.get(0)
                report.insert(0, 'Sample ID', report.pop('Sample ID'))
                report.insert(1, 'Contig ID', report.pop('Contig ID'))
                
                for col in report.columns:
                    labels = {}
                    df20PC = report.loc[report['Sequence Name'].
                                            str.contains('.20PC')]
                    index = 0
                    for col in df20PC.columns:

                        if col == 'Contig ID' : ###
                            df20PC = df20PC.rename(columns={col: 0})
                            labels[0] = col
                        if col != 'Contig ID' : ###
                            index += 1
                            df20PC = df20PC.rename(columns={col: (index + 0.2)})
                            labels[index + 0.2] = f'{col}_>20%'

                    df2PC = report.loc[report['Sequence Name'].str.contains('.2PC')]

                    index = 0
                    for col in df2PC.columns:

                        if col == 'Contig ID' : ###
                            df2PC = df2PC.rename(columns={col: 0})
                        if col != 'Contig ID' : ### Sample ID
                            index += 1
                            df2PC = df2PC.rename(columns={col: (index + 0.1)})
                            labels[index + 0.1] = f'{col}_2-20%'

                    SeqSummary20_2PC = pd.merge(df20PC, df2PC, on= 0) #on= 'Sample ID')
                    SeqSummary20_2PC = SeqSummary20_2PC[sorted(SeqSummary20_2PC)]
                    for n in SeqSummary20_2PC.columns:
                        if n != 0:
                            for key, value in labels.items():
                                if key == n:
                                    SeqSummary20_2PC = SeqSummary20_2PC.rename(columns={n: value})
                        if n == 0:
                            SeqSummary20_2PC = SeqSummary20_2PC.rename(columns={n: labels[0]})
                    SeqSummary20_2PC.rename(columns={'Contig ID_>20%': 
                                                    'Contig ID'}, inplace= True) ####

                'QC for Antiviral Resistance Genotyping, The filter is based only in 20PC'

                if title == 'SequenceSummary' or title == 'sequenceSummaries':
                    
                    SeqSummary20_2PC = SeqSummary20_2PC.drop(columns=[
                                    'Sequence Name_2-20%','Sequence Name_>20%',
                                    'Genes_2-20%','PR Start_2-20%', 'PR End_2-20%', 
                                    'RT Start_2-20%','RT End_2-20%', 'IN Start_2-20%', 
                                    'IN End_2-20%','Subtype (%)_2-20%', f'{Pcnt_Mix}_2-20%',
                                    ])
                    'FILTERS'
                    'Minumun coverage for each protein'
                    AAcoverage = (SeqSummary20_2PC['PR Start_>20%']<=30)\
                                & (SeqSummary20_2PC['PR End_>20%'] >=90)\
                                & (SeqSummary20_2PC['RT Start_>20%']<=65)\
                                & (SeqSummary20_2PC['RT End_>20%'] >=215)\
                                & (SeqSummary20_2PC['IN Start_>20%']<=66)\
                                & (SeqSummary20_2PC['IN End_>20%'] >=263)
                
                    '''Maximum number of APOBEC3G/F hypermutated AAs	
                    3 + 2 + 3 (RT, PR, IN, respectively)'''
                    apobec = SeqSummary20_2PC['Num Apobec Mutations_>20%'] <= 8

                    '''Maximum number of highly unusual AA mutations
                    15	8	10 (RT, PR, IN, respectively)'''
                    unusual = SeqSummary20_2PC['Num Unusual Mutations_>20%']<=33

                    '''Maximum number of stop codons
                    + unpublished AA insertions or deletions
                    + highly ambiguous nucelotides (B,D,H,V,N)
                    4	2	3 (RT, PR, IN, respectively)'''
                    bdhvn = SeqSummary20_2PC[f'{Num_BDHVN}_>20%']<=9
                
                    #'Once the results are filtered, presence of Ns in pol. 
                    # THIS CAN BE VERY STRINGENT!!!'
                    #NsInPRRTIN = (SeqSummary20_2PC['PR Other_>20%'].str.contains('X'))\
                    #            & (SeqSummary20_2PC['RT Other_>20%'].str.contains('X'))\
                    #            & (SeqSummary20_2PC['IN Other_>20%'].str.contains('X'))
                    good_QC_sequences = list(SeqSummary20_2PC.loc[AAcoverage & apobec & unusual
                                                            & bdhvn, 'Contig ID']) 
                                                            # & ~NsInPRRTIN, 'Sample ID'])

                    for n in range(len(SeqSummary20_2PC)):
                        if SeqSummary20_2PC.loc[n]['Contig ID'] in good_QC_sequences:
                            SeqSummary20_2PC.at[n ,'QC Stanford ARG'] = 'PASSED'
                            #SeqSummary20_2PC.loc[n ,'QC Stanford ARG'] = 'PASSED' # works similar
                        else:
                            SeqSummary20_2PC.at[n ,'QC Stanford ARG'] = 'NOT Passed'
                    
                    SeqSummary20_2PC.insert(2, 'QC Stanford ARG',
                                            SeqSummary20_2PC.pop('QC Stanford ARG'))

                    MutSummary = SeqSummary20_2PC[['Contig ID', 'QC Stanford ARG', 
                                                'PI SDRMs_2-20%', 'PI SDRMs_>20%',
                                                'PI Accessory_2-20%', 'PI Accessory_>20%', 
                                                'NRTI SDRMs_2-20%','NRTI SDRMs_>20%', 
                                                'NNRTI SDRMs_2-20%','NNRTI SDRMs_>20%', 
                                                'INSTI SDRMs_2-20%','INSTI SDRMs_>20%']]
  
                if title == 'ResistanceSummary' or title == 'resistanceSummaries' :
                    
                    DrugScoreSummary = SeqSummary20_2PC[[
                       'Contig ID', f'Genes_>20%', #'Sample ID',
                       'PI Major_2-20%','PI Major_>20%',
                       'ATV/r Level_2-20%','ATV/r Level_>20%',
                       'DRV/r Level_2-20%','DRV/r Level_>20%',
                       'FPV/r Level_2-20%','FPV/r Level_>20%',
                       'IDV/r Level_2-20%','IDV/r Level_>20%',
                       'LPV/r Level_2-20%','LPV/r Level_>20%',
                       'NFV Level_2-20%','NFV Level_>20%',
                       'SQV/r Level_2-20%','SQV/r Level_>20%',
                       'TPV/r Level_2-20%','TPV/r Level_>20%',
                       'NRTI_2-20%','NRTI_>20%',
                       'ABC Level_2-20%','ABC Level_>20%',
                       'AZT Level_2-20%','AZT Level_>20%',
                       'D4T Level_2-20%','D4T Level_>20%',
                       'DDI Level_2-20%','DDI Level_>20%',
                       'FTC Level_2-20%','FTC Level_>20%',
                       '3TC Level_2-20%','3TC Level_>20%',
                       'TDF Level_2-20%', 'TDF Level_>20%',
                       'NNRTI_2-20%','NNRTI_>20%',
                       'DOR Level_2-20%','DOR Level_>20%',
                       'EFV Level_2-20%','EFV Level_>20%',
                       'ETR Level_2-20%','ETR Level_>20%',
                       'NVP Level_2-20%','NVP Level_>20%',
                       'RPV Level_2-20%','RPV Level_>20%',
                       'INSTI Major_2-20%','INSTI Major_>20%',
                       'INSTI Accessory_2-20%','INSTI Accessory_>20%',
                       'BIC Level_2-20%','BIC Level_>20%',
                       'CAB Level_2-20%','CAB Level_>20%',
                       'DTG Level_2-20%','DTG Level_>20%',
                       'EVG Level_2-20%','EVG Level_>20%',
                       'RAL Level_2-20%','RAL Level_>20%',
                       'Algorithm Name_>20%','Algorithm Version_>20%',
                       'Algorithm Date_>20%']]
                    
            DrugScoreSummary = DrugScoreSummary.merge(MutSummary, how='left', on='Contig ID')
 
            DrugScoreSummary.rename(columns={'Genes_>20%': 'Genes', 
                                            'Algorithm Name_>20%': 'Algorithm Name',
                                            'Algorithm Version_>20%': 'Algorithm Version',
                                            'Algorithm Date_>20%': 'Algorithm Date'},
                                            inplace= True)
            DrugScoreSummary[['PI Frequency (%)','PI Depth','NRTI Frequency (%)', 
                            'NRTI Depth','NNRTI Frequency (%)',	'NNRTI Depth',
                            'INSTI Frequency (%)',	'INSTI Depth']] = ''

            ##################################################
            #n =0
            #for c in DrugScoreSummary.columns:
            #    print(c, '____',n)
            #    n+=1
            ################################################
            DrugScoreSummary.insert(52, 'INSTI Frequency (%)',
                                    DrugScoreSummary.pop('INSTI Frequency (%)'))
            DrugScoreSummary.insert(52, 'INSTI Depth',
                                    DrugScoreSummary.pop('INSTI Depth'))
            DrugScoreSummary.insert(48, 'INSTI SDRMs_>20%',
                                    DrugScoreSummary.pop('INSTI SDRMs_>20%'))
            DrugScoreSummary.insert(48, 'INSTI SDRMs_2-20%',
                                    DrugScoreSummary.pop('INSTI SDRMs_2-20%'))

            DrugScoreSummary.insert(38, 'NNRTI Frequency (%)',
                                    DrugScoreSummary.pop('NNRTI Frequency (%)'))
            DrugScoreSummary.insert(38, 'NNRTI Depth',
                                    DrugScoreSummary.pop('NNRTI Depth'))
            DrugScoreSummary.insert(36, 'NNRTI SDRMs_>20%',
                                    DrugScoreSummary.pop('NNRTI SDRMs_>20%'))
            DrugScoreSummary.insert(36, 'NNRTI SDRMs_2-20%',
                                    DrugScoreSummary.pop('NNRTI SDRMs_2-20%'))

            DrugScoreSummary.insert(22, 'NRTI Frequency (%)',
                                    DrugScoreSummary.pop('NRTI Frequency (%)'))
            DrugScoreSummary.insert(22, 'NRTI Depth',
                                    DrugScoreSummary.pop('NRTI Depth'))
            DrugScoreSummary.insert(20, 'NRTI SDRMs_>20%',
                                    DrugScoreSummary.pop('NRTI SDRMs_>20%'))
            DrugScoreSummary.insert(20, 'NRTI SDRMs_2-20%',
                                    DrugScoreSummary.pop('NRTI SDRMs_2-20%'))

            DrugScoreSummary.insert(4, 'PI Frequency (%)',
                                    DrugScoreSummary.pop('PI Frequency (%)'))
            DrugScoreSummary.insert(4, 'PI Depth',
                                    DrugScoreSummary.pop('PI Depth'))
            DrugScoreSummary.insert(4, 'PI Accessory_>20%',
                                    DrugScoreSummary.pop('PI Accessory_>20%'))
            DrugScoreSummary.insert(4, 'PI Accessory_2-20%',
                                    DrugScoreSummary.pop('PI Accessory_2-20%'))
            DrugScoreSummary.insert(2, 'PI SDRMs_>20%',
                                    DrugScoreSummary.pop('PI SDRMs_>20%'))
            DrugScoreSummary.insert(2, 'PI SDRMs_2-20%',
                                    DrugScoreSummary.pop('PI SDRMs_2-20%'))

            DrugScoreSummary.insert(1, 'QC Stanford ARG',
                                    DrugScoreSummary.pop('QC Stanford ARG'))

            sample_list = sample_list.merge(DrugScoreSummary, how='left', 
                                            left_on='Contig',
                                            right_on='Contig ID')

            sample_list.drop('Contig ID', axis=1, inplace=True)

            drugLevelCols = []
            for col in sample_list.columns:
                if 'Level' in col:
                    drugLevelCols.append(col)

            #df['col_new'] = df['col1'].map(dict)        
            resistanceCode = {  1: 'Susceptible',
                                2: 'Potential low-level resistance',
                                3: 'Low-level resistance',
                                4: 'Intermediate resistance',
                                5: 'High-level resistance'}
            for dlc in drugLevelCols:
                sample_list[dlc] = sample_list[dlc].map(resistanceCode)


            ''' As the threshold of the sequences 2PC includes everyhting 
            which is above 20PC this must be clean to group the mutations in
            2-20% and >20%'''
            # np.where(cond, if_cond_True, if_cond_False)
            'PI'
            sample_list['PI SDRMs_2-20%'] = np.where(
                            (sample_list['PI SDRMs_>20%'] != np.nan) & 
                            (sample_list['PI SDRMs_2-20%'] != np.nan) &
                            (sample_list['PI SDRMs_2-20%'] == 
                                        sample_list['PI SDRMs_>20%']) , 
                            np.nan,
                            sample_list['PI SDRMs_2-20%'])

            sample_list['PI Major_2-20%'] = np.where(
                            (sample_list['PI Major_>20%'] != np.nan) & 
                            (sample_list['PI Major_2-20%'] != np.nan) &
                            (sample_list['PI Major_>20%'] != 'None') & 
                            (sample_list['PI Major_2-20%'] != 'None') &
                            (sample_list['PI Major_2-20%'] == 
                                        sample_list['PI Major_>20%']) , 
                            'None',
                            sample_list['PI Major_2-20%']
                            )

            sample_list['PI Accessory_2-20%'] = np.where(
                            (sample_list['PI Accessory_>20%'] != np.nan) & 
                            (sample_list['PI Accessory_2-20%'] != np.nan) &
                            (sample_list['PI Accessory_>20%'] != 'None') & 
                            (sample_list['PI Accessory_2-20%'] != 'None') &
                            (sample_list['PI Accessory_2-20%'] == 
                                        sample_list['PI Accessory_>20%']) , 
                            'None',
                            sample_list['PI Accessory_2-20%']
                            )

            'NRTI 2-20%'
            sample_list['NRTI SDRMs_2-20%'] = np.where(
                            (sample_list['NRTI SDRMs_>20%'] != np.nan) & 
                            (sample_list['NRTI SDRMs_2-20%'] != np.nan) &
                            (sample_list['NRTI SDRMs_2-20%'] == 
                                        sample_list['NRTI SDRMs_>20%']) , 
                            np.nan,
                            sample_list['NRTI SDRMs_2-20%'])

            sample_list['NRTI_2-20%'] = np.where(
                            (sample_list['NRTI_>20%'] != np.nan) & 
                            (sample_list['NRTI_2-20%'] != np.nan) &
                            (sample_list['NRTI_>20%'] != 'None') & 
                            (sample_list['NRTI_2-20%'] != 'None') &
                            (sample_list['NRTI_2-20%'] == 
                                        sample_list['NRTI_>20%']) , 
                            'None',
                            sample_list['NRTI_2-20%']
                            )

            'NNRTI 2-20%'
            sample_list['NNRTI SDRMs_2-20%'] = np.where(
                            (sample_list['NNRTI SDRMs_>20%'] != np.nan) & 
                            (sample_list['NNRTI SDRMs_2-20%'] != np.nan) &
                            (sample_list['NNRTI SDRMs_2-20%'] == 
                                        sample_list['NNRTI SDRMs_>20%']) , 
                            np.nan,
                            sample_list['NNRTI SDRMs_2-20%'])

            sample_list['NNRTI_2-20%'] = np.where(
                            (sample_list['NNRTI_>20%'] != np.nan) & 
                            (sample_list['NNRTI_2-20%'] != np.nan) &
                            (sample_list['NNRTI_>20%'] != 'None') & 
                            (sample_list['NNRTI_2-20%'] != 'None') &
                            (sample_list['NNRTI_2-20%'] == 
                                        sample_list['NNRTI_>20%']) , 
                            'None',
                            sample_list['NNRTI_2-20%']
                            )

            'INSTI'
            sample_list['INSTI SDRMs_2-20%'] = np.where(
                            (sample_list['INSTI SDRMs_>20%'] != np.nan) & 
                            (sample_list['INSTI SDRMs_2-20%'] != np.nan) &
                            (sample_list['INSTI SDRMs_2-20%'] == 
                                        sample_list['INSTI SDRMs_>20%']) , 
                            np.nan,
                            sample_list['INSTI SDRMs_2-20%'])

            sample_list['INSTI Major_2-20%'] = np.where(
                            (sample_list['INSTI Major_>20%'] != np.nan) & 
                            (sample_list['INSTI Major_2-20%'] != np.nan) &
                            (sample_list['INSTI Major_>20%'] != 'None') & 
                            (sample_list['INSTI Major_2-20%'] != 'None') &
                            (sample_list['INSTI Major_2-20%'] == 
                                        sample_list['INSTI Major_>20%']) , 
                            'None',
                            sample_list['INSTI Major_2-20%']
                            )

            sample_list['INSTI Accessory_2-20%'] = np.where(
                            (sample_list['INSTI Accessory_>20%'] != np.nan) & 
                            (sample_list['INSTI Accessory_2-20%'] != np.nan) &
                            (sample_list['INSTI Accessory_>20%'] != 'None') & 
                            (sample_list['INSTI Accessory_2-20%'] != 'None') &
                            (sample_list['INSTI Accessory_2-20%'] == 
                                        sample_list['INSTI Accessory_>20%']) , 
                            'None',
                            sample_list['INSTI Accessory_2-20%']
                            )

            os.chdir(run_path)
            sample_list.to_csv(f'{run_ID}_Sequencing_results.csv', index= False)

            os.chdir(batch_path)
            excel_file_name = Path(f'{batch}_NGS_Results_2-20PC_D100.xlsx')
            previous_runs = []
            if Path(excel_file_name).exists():
                wb = load_workbook(excel_file_name)
                ws1 = wb['Summary_2-20PC_Depth_100']
                for row in ws1.iter_cols(min_col= 2, max_col=2, 
                            values_only= True):
                    for cell_value in row:
                        if (cell_value not in previous_runs 
                                and cell_value != 'NGS run'):
                            previous_runs.append(cell_value.upper())
                previous_runs = list(set(previous_runs))
                if run_ID.upper() not in previous_runs:
                    for rowdf in dataframe_to_rows(sample_list, index= False, 
                                                   header= False):
                        ws1.append(rowdf)
                for cell in ws1[1]:
                    cell.style = 'Accent1'
                wb.save(excel_file_name)

            else:
                wb = Workbook()
                #ws = wb.active
                ws1 = wb.create_sheet('Summary_2-20PC_Depth_100', 0)
                ws2 = wb.create_sheet('QC Stanford ARG Criteria', 1)

                criteria_row1 = 'Minimum amino acid(AA) coverage'
                criteria_row2 = 'Maximum number of stop codons + unpublished\
                                AA insertions or deletions + highly ambiguous\
                                nucelotides (B,D,H,V,N)'
                criteria_row3 = 'Maximum number of APOBEC3G/F hypermutated AAs'
                criteria_row4 = 'Maximum number of highly unusual AA mutations'
                criteria = pd.DataFrame({ 'Criteria':[criteria_row1, 
                                                      criteria_row2, 
                                                      criteria_row3,
                                                      criteria_row4],
                                        'RT':['Positions 65 to 215', 4, 3, 15],
                                        'PR':['Positions 30 to 90', 2, 2, 8],
                                        'IN':['Positions 66 to 263', 3,3 ,10]
                                        })


                for rowdf in dataframe_to_rows(criteria, index= False, header= True):
                    ws2.append(rowdf)
                for rowdf in dataframe_to_rows(sample_list, index= False, 
                                                       header= True):
                    ws1.append(rowdf)

                for cell in ws2[1]:
                    cell.style = 'Accent2' #'Pandas'
                for cell in ws1[1]:
                    cell.style = 'Accent1'
                    wb.save(excel_file_name)
            logging.info('Antiviral resistance genotyping data added')
            fasta_batches(run_ID, sample_list)

        else:

            logging.warning(f'''
    < STANFORD REPORT WARNING >
    A folder "{run_ID}" already exists in "{run_ID}/hivdb.stanford.report".
    It is possible that the current run had been previously analysed and 
    the data already added to the file "{batch}_NGS_Results_2-20PC_D100.xlsx"
    in the folder "{batch_path.stem}".
        ''')
            sys.exit(1)    
    else:
        logging.warning(f'''
    < STANFORD REPORT FILE ERROR >
    Confirm that a file called "{run_ID}.zip" exists in "{run_ID}/hivdb.stanford.report".
                ''')
        sys.exit(1)

    #"""#
##############################################################################

def majority_d30(run_ID, batch, batch_path):

    """ Consolidates all the sample information about the subtyping and resistance
    genotyping using sequences at depth of 30 reads and frequency of 20PC
    (Majority Variants, respectively) in an XLSX file in the main 
    directory of the INITIO batch."""

    
    Depth30 = '_D30'
    sample_list = sample_info(run_ID)
    sample_list = sequencing_metrics(sample_list)
    sample_list = subtyping_data(run_ID, sample_list, Depth= Depth30)
 
    ''' Antiviral resistance genotyping D30'''
    
    stanford_results = Path.cwd().joinpath('hivdb.stanford.report')
    os.chdir(stanford_results)

    if stanford_results.joinpath(f'{run_ID}{Depth30}.zip').is_file(): 

        if Path.cwd().joinpath(f'{run_ID}{Depth30}').exists() == False: ##: #

            sp.run(['unzip','-j', f'{run_ID}{Depth30}.zip', 
                    '-d', f'{run_ID}{Depth30}'])#, text=True)

            try:

                SequenceSummary = pd.read_csv(
                        Path(f'{run_ID}{Depth30}').joinpath('SequenceSummary.tsv'), 
                        sep = '\t')
                ResistanceSummary = pd.read_csv(
                        Path(f'{run_ID}{Depth30}').joinpath('ResistanceSummary.tsv')
                        , sep = '\t')
                csv = { 'SequenceSummary': SequenceSummary, 
                        'ResistanceSummary': ResistanceSummary}
                Pcnt_Mix = 'Pcnt Mix'
                StopCodons = 'StopCodons'
                Num_BDHVN = 'Num BDHVN'
                BDHVN = 'BDHVN' 
                UnusualMutations = 'UnusualMutations'
                
            except:

                SequenceSummary = pd.read_csv(
                        Path(f'{run_ID}{Depth30}').joinpath('sequenceSummaries.csv'))
                ResistanceSummary = pd.read_csv(
                        Path(f'{run_ID}{Depth30}').joinpath('resistanceSummaries.csv'))
                csv = { 'sequenceSummaries': SequenceSummary, 
                        'resistanceSummaries': ResistanceSummary}
                Pcnt_Mix = 'NA Mixture Rate (%)'
                StopCodons = 'Stop Codons'
                Num_BDHVN = 'Num Ambiguous'
                BDHVN = 'Ambiguous'
                UnusualMutations = 'Unusual Mutations'
                
            MutSummary = pd.DataFrame()
            DrugScoreSummary = pd.DataFrame()
            for title, report in csv.items():
                report['Contig ID'] = report['Sequence Name'].\
                                        str.replace(r'\.(2|20)PC.*', '' , 
                                        regex= True)
                report['Sample ID'] = report['Sequence Name'].\
                                        str.split('.').\
                                        str.get(0)
                report.insert(0, 'Sample ID', report.pop('Sample ID'))
                report.insert(1, 'Contig ID', report.pop('Contig ID'))

                'QC for Antiviral Resistance Genotyping, The filter is based only in 20PC'

                if title == 'SequenceSummary' or title == 'sequenceSummaries':
                    
                    'FILTERS'
                    'Minumun coverage for each protein'
                    AAcoverage = (report['PR Start']<=30)\
                            & (report['PR End'] >=90)\
                            & (report['RT Start']<=65)\
                            & (report['RT End'] >=215)\
                            & (report['IN Start']<=66)\
                            & (report['IN End'] >=263)
                
                    '''Maximum number of APOBEC3G/F hypermutated AAs	
                    3 + 2 + 3 (RT, PR, IN, respectively)'''
                    apobec = report['Num Apobec Mutations'] <= 8

                    '''Maximum number of highly unusual AA mutations
                    15	8	10 (RT, PR, IN, respectively)'''
                    unusual = report['Num Unusual Mutations']<=33

                    '''Maximum number of stop codons
                    + unpublished AA insertions or deletions
                    + highly ambiguous nucelotides (B,D,H,V,N)
                    4	2	3 (RT, PR, IN, respectively)'''
                    bdhvn = report[f'{Num_BDHVN}']<=9
                
                    #'Once the results are filtered, presence of Ns in pol. 
                    # THIS CAN BE VERY STRINGENT!!!'
                    #NsInPRRTIN = (report['PR Other'].str.contains('X'))\
                    #            & (report['RT Other'].str.contains('X'))\
                    #            & (report['IN Other'].str.contains('X'))
                    good_QC_sequences = list(report.loc[AAcoverage & apobec & unusual
                                                            & bdhvn, 'Contig ID']) 
                                                            # & ~NsInPRRTIN, 'Sample ID'])
                    for n in range(len(report)):
                        if report.loc[n]['Contig ID'] in good_QC_sequences:
                            report.at[n ,'QC Stanford ARG'] = 'PASSED'
                            #SeqSummary20_2PC.loc[n ,'QC Stanford ARG'] = 'PASSED' 
                            # # works similar
                        else:
                            report.at[n ,'QC Stanford ARG'] = 'NOT Passed'
                    report.insert(2, 'QC Stanford ARG',
                                            report.pop('QC Stanford ARG'))

                    MutSummary = report[['Contig ID', 'QC Stanford ARG', 
                                    'PI SDRMs','PI Accessory', 
                                    'NRTI SDRMs', 'NNRTI SDRMs', 
                                    'INSTI SDRMs']]
  
                if title == 'ResistanceSummary' or title == 'resistanceSummaries' :
                    
                    DrugScoreSummary = report[[
                        'Contig ID', 'Genes', #'Sample ID',
                        'PI Major','ATV/r Level','DRV/r Level','FPV/r Level',
                        'IDV/r Level','LPV/r Level','NFV Level','SQV/r Level',
                        'TPV/r Level',
                        'NRTI','ABC Level','AZT Level','D4T Level','DDI Level',
                        'FTC Level','3TC Level','TDF Level',
                        'NNRTI','DOR Level','EFV Level','ETR Level','NVP Level',
                        'RPV Level',
                        'INSTI Major','INSTI Accessory','BIC Level','CAB Level',
                        'DTG Level','EVG Level','RAL Level',
                        'Algorithm Name','Algorithm Version','Algorithm Date']]
                    
            DrugScoreSummary = DrugScoreSummary.merge(MutSummary, how='left', on='Contig ID')
 
            DrugScoreSummary[['PI Frequency (%)','PI Depth','NRTI Frequency (%)', 
                        'NRTI Depth','NNRTI Frequency (%)',	'NNRTI Depth',
                        'INSTI Frequency (%)',	'INSTI Depth']] = ''

            ##################################################
            #n =0
            #for c in DrugScoreSummary.columns:
            #    print(c, '____',n)
            #    n+=1
            ################################################
            DrugScoreSummary.insert(27, 'INSTI Frequency (%)',
                                    DrugScoreSummary.pop('INSTI Frequency (%)'))
            DrugScoreSummary.insert(27, 'INSTI Depth',
                                    DrugScoreSummary.pop('INSTI Depth'))
            DrugScoreSummary.insert(25, 'INSTI SDRMs',
                                    DrugScoreSummary.pop('INSTI SDRMs'))

            DrugScoreSummary.insert(20, 'NNRTI Frequency (%)',
                                    DrugScoreSummary.pop('NNRTI Frequency (%)'))
            DrugScoreSummary.insert(20, 'NNRTI Depth',
                                    DrugScoreSummary.pop('NNRTI Depth'))
            DrugScoreSummary.insert(19, 'NNRTI SDRMs',
                                    DrugScoreSummary.pop('NNRTI SDRMs'))

            DrugScoreSummary.insert(12, 'NRTI Frequency (%)',
                                    DrugScoreSummary.pop('NRTI Frequency (%)'))
            DrugScoreSummary.insert(12, 'NRTI Depth',
                                    DrugScoreSummary.pop('NRTI Depth'))
            DrugScoreSummary.insert(11, 'NRTI SDRMs',
                                    DrugScoreSummary.pop('NRTI SDRMs'))

            DrugScoreSummary.insert(3, 'PI Frequency (%)',
                                    DrugScoreSummary.pop('PI Frequency (%)'))
            DrugScoreSummary.insert(3, 'PI Depth',
                                    DrugScoreSummary.pop('PI Depth'))
            DrugScoreSummary.insert(3, 'PI Accessory',
                                    DrugScoreSummary.pop('PI Accessory'))
            DrugScoreSummary.insert(2, 'PI SDRMs',
                                    DrugScoreSummary.pop('PI SDRMs'))

            DrugScoreSummary.insert(1, 'QC Stanford ARG',
                                    DrugScoreSummary.pop('QC Stanford ARG'))

            sample_list = sample_list.merge(DrugScoreSummary, how='left', 
                                            left_on='Contig',
                                            right_on='Contig ID')

            sample_list.drop('Contig ID', axis=1, inplace=True)

            drugLevelCols = []
            for col in sample_list.columns:
                if 'Level' in col:
                    drugLevelCols.append(col)

            #df['col_new'] = df['col1'].map(dict)        
            resistanceCode = {  1: 'Susceptible',
                                2: 'Potential low-level resistance',
                                3: 'Low-level resistance',
                                4: 'Intermediate resistance',
                                5: 'High-level resistance'}
            for dlc in drugLevelCols:
                sample_list[dlc] = sample_list[dlc].map(resistanceCode)

            os.chdir(run_path)
            sample_list.to_csv(
                    f'{run_ID}_Sequencing_results_Majority_Variants{Depth30}.csv', 
                    index= False
                    )

            os.chdir(batch_path)
            excel_file_name = Path(
                        f'{batch}_NGS_Results_Majority_Variants{Depth30}.xlsx')
            previous_runs = []
            if Path(excel_file_name).exists():
                wb = load_workbook(excel_file_name)
                ws1 = wb['Summary_Depth_30_20PC']
                for row in ws1.iter_cols(min_col= 2, max_col=2, 
                            values_only= True):
                    for cell_value in row:
                        if (cell_value not in previous_runs 
                                and cell_value != 'NGS run'):
                            previous_runs.append(cell_value.upper())
                previous_runs = list(set(previous_runs))
                if run_ID.upper() not in previous_runs:
                    for rowdf in dataframe_to_rows(sample_list, index= False, 
                                                   header= False):
                        ws1.append(rowdf)
                for cell in ws1[1]:
                    cell.style = 'Accent1'
                wb.save(excel_file_name)

            else:
                wb = Workbook()
                #ws = wb.active
                ws1 = wb.create_sheet('Summary_Depth_30_20PC', 0)
                ws2 = wb.create_sheet('QC Stanford ARG Criteria', 1)
                
                criteria_row1 = 'Minimum amino acid(AA) coverage'
                criteria_row2 = 'Maximum number of stop codons + unpublished\
                                AA insertions or deletions + highly ambiguous\
                                nucelotides (B,D,H,V,N)'
                criteria_row3 = 'Maximum number of APOBEC3G/F hypermutated AAs'
                criteria_row4 = 'Maximum number of highly unusual AA mutations'
                criteria = pd.DataFrame({ 'Criteria':[criteria_row1, 
                                                      criteria_row2, 
                                                      criteria_row3,
                                                      criteria_row4],
                                        'RT':['Positions 65 to 215', 4, 3, 15],
                                        'PR':['Positions 30 to 90', 2, 2, 8],
                                        'IN':['Positions 66 to 263', 3,3 ,10]
                                        })


                for rowdf in dataframe_to_rows(criteria, index= False, 
                                header= True):
                    ws2.append(rowdf)
                for rowdf in dataframe_to_rows(sample_list, index= False, 
                                header= True):
                    ws1.append(rowdf)

                for cell in ws2[1]:
                    cell.style = 'Accent2' #'Pandas'
                for cell in ws1[1]:
                    cell.style = 'Accent1'
                    wb.save(excel_file_name)

            fasta_batches(run_ID, sample_list, depth= 30)

        else:

            logging.warning(f'''
    < STANFORD REPORT WARNING >
    A folder "{run_ID}{Depth30}" already exists in "{run_ID}/hivdb.stanford.report".
    It is possible that the current run had been previously analysed and 
    the data already added to the file "{batch}_NGS_Results_Majority_Variants{Depth30}.xlsx"
    in the folder "{batch_path.stem}".
        ''')
            sys.exit(1)    
    else:
        logging.warning(f'''
    < STANFORD REPORT FILE ERROR >
    Confirm that a file called "{run_ID}{Depth30}.zip" exist in "{run_ID}/hivdb.stanford.report".
                ''')
        sys.exit(1)

##############################################################################


if __name__ == "__main__":
    
    logging.basicConfig(format='%(asctime)s %(levelname)s: %(message)s',
                            datefmt='%Y-%m-%d %I:%M',
                            level=logging.INFO)
    
    if Path(args.input).is_dir():

        run_info = args.input

        run_path, run_ID, batch_path, \
        batch, project, ref_seq, \
        quasibams, data_to_clean, tmps = parsing_paths(run_info)

        logging.info(f'Processing data from directory {run_ID}')

        if args.post_genomancer:
            depth = args.post_genomancer.upper()
            logging.info(f'Analysis on Genomancer results for {run_ID}')
            postgenomancer(run_ID, depth)
            logging.info(f'Process successfully COMPLETE {run_ID}')

        elif args.quality_control:
            logging.info(f'Quality Control (QC) for sequences {run_ID}')
            qc_for_sequence_analysis(quasibams, ref_seq, data_to_clean, tmps)
            logging.info(f'QC successfully COMPLETE {run_ID}')

        elif args.confirmation:
            logging.info(f'Confirmation of quality for sequences {run_ID}')
            confirmation_of_sequence_cleaning(run_ID, ref_seq, quasibams, 
                                            data_to_clean, tmps)
            logging.info(f'Process successfully COMPLETE for {run_ID}')

        elif args.file_generation:
            logging.info(f'Generation of FASTA files for data analysis {run_ID}')
            file_generation_for_data_analysis(run_ID, ref_seq, quasibams, 
                                            data_to_clean, tmps)
            logging.info('File generation successfully COMPLETE for {run_ID}')
        
        elif args.post_jphmm:
            jphmm_report = args.post_jphmm
            logging.info('Processing jphmm subtyping results for {run_ID}')
            process_jphmm_results(jphmm_report,  run_ID)
            logging.info('Process succesfully COMPLETE for {run_ID}')

        elif args.data_consolidation:
            logging.info(f'Consolidation of data for {run_ID}, variants at 2 and 20% frequency, depth 100 reads')
            data_consolidator(run_ID, batch, batch_path)
            logging.info(f'Data consolidation for {run_ID} successfully COMPLETE.')
            logging.info(f'Check the directory "{batch}/FASTA_sequences" for new updates in the FASTA files.')
            logging.info(f'Check the file "{batch}_NGS_Results_2-20PC_D100.xlsx" to see the new data.')

        elif args.majority_30:
            logging.info(f'Consolidation of data for {run_ID}, majority variants, depth 30 reads')
            majority_d30(run_ID, batch, batch_path)
            logging.info(f'Data consolidation for {run_ID} successfully COMPLETE.')
            logging.info(f'Check the directory "{batch}/FASTA_sequences" for new updates in the FASTA files.')
            logging.info(f'Check the file "{batch}_NGS_Results_Majority_Variants_Depth_30.xlsx" to see the new data.')

        else:
            logging.info(f'{program_specs.print_help()}')
            logging.warning(f'''

Please enter --input and one of the following argument according to the step to carry out:

1) Analysis of data after using pipeline genomancer: -pgen, --post_genomancer 
2) Quality control: -qc, --quality_control
3) Confirmation of the sequences: -c, --confirmation
4) Generation of files to perform data analysis (subtyping, resistance..): -fg, --file_generation
5) Run jumping profile Hidden Markov Model for HIV locally:-j, --jphmm
6) Generate a summary with the results from jpHHM:-pj, --post_jphmm
7) Consolidation of data (2/20PC depth 100reads) from subtyping, resistance, ... and generation of final report: -dc, --data_consolidation
8) Consolidation of data for Majority Variants at 30reads: -majodc, --majority_30 

If needed check the help displayed above. 

Cheers!
''')


    if Path(args.input).is_file():

        run_info = args.input
        file_name = Path(run_info).name
        logging.info(f'> Processing data from file {file_name}')
    
        if args.jphmm:
            logging.info('Subtyping by jpHMM in PROCESS {run_ID}')
            running_local_jphmm(run_info)
            logging.info('Subtyping complete for {run_ID}')

        else:
            logging.warning(f'''
Please select a specific FASTA file to analyse using --input 
and enter the argument -j or --jphmm
Cheers!
    ''')

#"""
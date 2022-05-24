#!/usr/bin/env python3
import argparse
import subprocess as sp
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import re

''' This script works to filter the antiviral resistance genpotyping data
taking as argument SequenceSummary.tsv to filter the sequences with good 
quality'''

parser = argparse.ArgumentParser(description='it takes the path to the NGS run folder')
parser.add_argument('--input', '-i', required=True) #INITIO_TESTING/Run01_NGS51/
arg = parser.parse_args()

path_to_rundir = Path.cwd().joinpath(arg.input)
run_id = path_to_rundir.stem 
initio_batch = path_to_rundir.parts[-2]
hiv_reports = Path.cwd().parents[0]/initio_batch/ f'hivdb.stanford.reports'
os.chdir(hiv_reports)

if Path(run_id).exists() == False:
    try:
        sp.run(['unzip','-j', f'{run_id}.zip', '-d', f'{run_id}'])#, text=True)

        SequenceSummary_file = Path(run_id).joinpath('SequenceSummary.tsv')
        ResistanceSummary_file = Path(run_id).joinpath('ResistanceSummary.tsv')
        SequenceSummary = pd.read_csv(SequenceSummary_file, sep = '\t')
        ResistanceSummary = pd.read_csv(ResistanceSummary_file, sep = '\t')
        csv = { 'SequenceSummary': SequenceSummary, 
                'ResistanceSummary': ResistanceSummary}
        MutSummary = pd.DataFrame()
        excel_file_name = Path(f'{initio_batch}_Stanford_Summary_2-20PC.xlsx')

        previous_runs = []
        if Path(excel_file_name).exists():
            
            wb = load_workbook(excel_file_name)
            ws1 = wb['MutationSummary_2-20PC']
            ws2 = wb['SequenceSummary_2-20PC']
            ws3 = wb['ResistanceSummary_2-20PC']
            for row in ws1.iter_cols(min_col= 2, max_col=2, values_only= True):
                for cell_value in row:
                    if (cell_value not in previous_runs 
                            and cell_value != 'NGS run'):
                        previous_runs.append(cell_value)
        else:
            wb = Workbook()
            #ws = wb.active
            ws1 = wb.create_sheet('MutationSummary_2-20PC', 0)
            ws2 = wb.create_sheet('SequenceSummary_2-20PC', 1)
            ws3 = wb.create_sheet('ResistanceSummary_2-20PC', 2)
            ws4 = wb.create_sheet('Criteria QC AntiResisGenotyping', 3)
            criteria = pd.DataFrame({ 'Criteria':['Minimum amino acid(AA) coverage',
                        'Maximum number of stop codons + unpublished AA insertions or deletions + highly ambiguous nucelotides (B,D,H,V,N)',
                            'Maximum number of APOBEC3G/F hypermutated AAs', 'Maximum number of highly unusual AA mutations'],
                                        'RT':['Positions 65 to 215', 4, 3, 15],
                                        'PR':['Positions 30 to 90', 2, 2, 8],
                                        'IN':['Positions 66 to 263', 3,3 ,10]})
            for rowdf in dataframe_to_rows(criteria, index= False, header= True):
                        ws4.append(rowdf)

        if run_id not in previous_runs:
            for title, report in csv.items():
                report['Sequence ID'] = report['Sequence Name'].\
                                            str.split('_').\
                                            str.get(0)
                report['Sample ID'] = report['Sequence Name'].\
                                            str.split('.').\
                                            str.get(0)

                report.insert(0, 'Sample ID', report.pop('Sample ID'))
                report.insert(1, 'Sequence ID', report.pop('Sequence ID'))
                for col in report.columns:
                    labels = {}
                    df20PC = report.loc[report['Sequence Name'].str.contains('_20PC')]
                    index = 0
                    for col in df20PC.columns:
                        if col == 'Sample ID' :
                            df20PC = df20PC.rename(columns={col: 0})
                            labels[0] = col
                        if col != 'Sample ID' :
                            index += 1
                            df20PC = df20PC.rename(columns={col: (index + 0.2)})
                            labels[index + 0.2] = f'{col}_20%'

                    df2PC = report.loc[report['Sequence Name'].str.contains('_2PC')]
                    index = 0
                    for col in df2PC.columns:
                        if col == 'Sample ID' :
                            df2PC = df2PC.rename(columns={col: 0})
                        if col != 'Sample ID' :
                            index += 1
                            df2PC = df2PC.rename(columns={col: (index + 0.1)})
                            labels[index + 0.1] = f'{col}_2%'
                    SeqSummary20_2PC = pd.merge(df20PC, df2PC, on= 0) #on= 'Sample ID')
                    SeqSummary20_2PC = SeqSummary20_2PC[sorted(SeqSummary20_2PC)]

                    for n in SeqSummary20_2PC.columns:
                        if n != 0:
                            for key, value in labels.items():
                                if key == n:
                                    SeqSummary20_2PC = SeqSummary20_2PC.rename(columns={n: value})
                        if n == 0:
                            SeqSummary20_2PC = SeqSummary20_2PC.rename(columns={n: labels[0]})

                    SeqSummary20_2PC['NGS run'] = run_id
                    SeqSummary20_2PC.insert(2, 'NGS run', SeqSummary20_2PC.pop('NGS run'))
                    SeqSummary20_2PC.rename(columns={'Sequence ID_20%': 'Sequence ID'}, inplace= True)

                'QC for Antiviral Resistance Genotyping, The filter is based only in 20PC'

                if title == 'SequenceSummary':
                    SeqSummary20_2PC = SeqSummary20_2PC.drop(columns=[
                                        'Sequence Name_2%','Sequence Name_20%',
                                        'Genes_2%','PR Start_2%', 'PR End_2%', 
                                        'RT Start_2%','RT End_2%', 'IN Start_2%', 
                                        'IN End_2%','Subtype (%)_2%', 'Pcnt Mix_2%',
                                        'Sequence ID_2%' ])
                    'FILTERS'
                    'Minumun coverage for each protein'
                    AAcoverage = (SeqSummary20_2PC['PR Start_20%']<=30)\
                                & (SeqSummary20_2PC['PR End_20%'] >=90)\
                                & (SeqSummary20_2PC['RT Start_20%']<=65)\
                                & (SeqSummary20_2PC['RT End_20%'] >=215)\
                                & (SeqSummary20_2PC['IN Start_20%']<=66)\
                                & (SeqSummary20_2PC['IN End_20%'] >=263)

                    '''Maximum number of APOBEC3G/F hypermutated AAs	
                    3 + 2 + 3 (RT, PR, IN, respectively)'''
                    apobec = SeqSummary20_2PC['Num Apobec Mutations_20%'] <= 8

                    '''Maximum number of highly unusual AA mutations
                    15	8	10 (RT, PR, IN, respectively)'''
                    unusual = SeqSummary20_2PC['Num Unusual Mutations_20%']<=33

                    '''Maximum number of stop codons
                    + unpublished AA insertions or deletions
                    + highly ambiguous nucelotides (B,D,H,V,N)
                    4	2	3 (RT, PR, IN, respectively)'''
                    bdhvn = SeqSummary20_2PC['Num BDHVN_20%']<=9

                    #'Once the results are filtered, presence of Ns in pol. 
                    # THIS CAN BE VERY STRINGENT!!!'
                    #NsInPRRTIN = (SeqSummary20_2PC['PR Other_20%'].str.contains('X'))\
                    #            & (SeqSummary20_2PC['RT Other_20%'].str.contains('X'))\
                    #            & (SeqSummary20_2PC['IN Other_20%'].str.contains('X'))

                    good_QC_sequences = list(SeqSummary20_2PC.loc[AAcoverage & apobec & unusual
                                                            & bdhvn, 'Sample ID']) # & ~NsInPRRTIN, 'Sample ID'])

                    for n in range(len(SeqSummary20_2PC)):
                        if SeqSummary20_2PC.loc[n]['Sample ID'] in good_QC_sequences:
                            SeqSummary20_2PC.at[n ,'QC Stanford ARG'] = 'PASSED'
                            #SeqSummary20_2PC.loc[n ,'QC Stanford ARG'] = 'PASSED' # works similar
                        else:
                            SeqSummary20_2PC.at[n ,'QC Stanford ARG'] = 'NOT Passed'

                    SeqSummary20_2PC.insert(2, 'QC Stanford ARG',
                                            SeqSummary20_2PC.pop('QC Stanford ARG'))

                    'Update the info in excel file '

                    #if os.path.isfile(excel_file_name):
                    if Path(excel_file_name).is_file():
                        for rowdf in dataframe_to_rows(SeqSummary20_2PC, 
                                                       index= False, 
                                                       header= False):
                            ws2.append(rowdf)
                    else:
                        for rowdf in dataframe_to_rows(SeqSummary20_2PC, 
                                                       index= False, 
                                                       header= True):
                            ws2.append(rowdf)
                    for cell in ws2[1]:
                        cell.style = 'Accent2' #'Pandas'

                    SeqSummary20_2PC['NGS run'] = run_id
                    MutSummary = SeqSummary20_2PC[['Sample ID','Sequence ID', 
                                                'NGS run','QC Stanford ARG', 
                                                'Genes_20%','PI SDRMs_2%', 
                                                'PI SDRMs_20%','PI Major_2%', 
                                                'PI Major_20%','PI Accessory_2%', 
                                                'PI Accessory_20%', 'NRTI SDRMs_2%',
                                                'NRTI SDRMs_20%', 'NRTI_2%', 
                                                'NRTI_20%', 'NNRTI SDRMs_2%',
                                                'NNRTI SDRMs_20%', 'NNRTI_2%',
                                                'NNRTI_20%', 'INSTI SDRMs_2%',
                                                'INSTI SDRMs_20%', 'INSTI Major_2%', 
                                                'INSTI Major_20%','INSTI Accessory_2%', 
                                                'INSTI Accessory_20%','IN Other_2%',
                                                'IN Other_20%']]

                if title == 'ResistanceSummary':
                    SeqSummary20_2PC = SeqSummary20_2PC.drop(columns=[
                                                'Sequence Name_2%','Sequence Name_20%',
                                                'Strain_2%','Strain_20%','Genes_2%',
                                                'Algorithm Name_2%', 'Algorithm Version_2%',
                                                'Algorithm Date_2%', 'Sequence ID_2%'])
                    for n in range(len(SeqSummary20_2PC)):
                        if SeqSummary20_2PC.loc[n]['Sample ID'] in good_QC_sequences:
                            SeqSummary20_2PC.at[n ,'QC Stanford ARG'] = 'PASSED'
                        else:
                            SeqSummary20_2PC.at[n ,'QC Stanford ARG'] = 'NOT Passed'

                    SeqSummary20_2PC.insert(2, 'QC Stanford ARG',
                                            SeqSummary20_2PC.pop('QC Stanford ARG'))
                    MutSummary['HIVDB Algorithm Version'] = SeqSummary20_2PC['Algorithm Version_20%']
                    MutSummary['HIVDB Algorithm Date'] = SeqSummary20_2PC['Algorithm Date_20%']

                    #if os.path.isfile(excel_file_name):
                    if Path(excel_file_name).is_file():
                        for rowdf in dataframe_to_rows(SeqSummary20_2PC, 
                                                       index= False, 
                                                       header= False):
                            ws3.append(rowdf)
                    else:
                        for rowdf in dataframe_to_rows(SeqSummary20_2PC, 
                                                       index= False, 
                                                       header= True):
                            ws3.append(rowdf)
                    for cell in ws3[1]:
                        cell.style = 'Accent3' #'Pandas'

            #if os.path.isfile(excel_file_name):
            if Path(excel_file_name).is_file():
                for rowdf in dataframe_to_rows(MutSummary, index= False, 
                                               header= False):
                    ws1.append(rowdf)
            else:
                for rowdf in dataframe_to_rows(MutSummary, index= False, 
                                               header= True):
                    ws1.append(rowdf)
            for cell in ws1[1]:
                        cell.style = 'Accent1'

            wb.save(excel_file_name)
        else:
            xl_name = excel_file_name.split('/')[-1]
            print(f'Data from {run_id} has been previously updated in the file {xl_name}')
        #os.remove(f'{run_id}.zip')

    except:
           
        print(f'''
    Make sure the file {run_id}.zip has been generated and saved in 
    
    {hiv_reports}
 
    or the report zip file has similar name to the run name {run_id}.
    ''')

else:
    print(f'The folder {run_id} already exist in {Path.cwd()}')
